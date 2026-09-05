"""Approved curated-data preview and export workflows."""

from __future__ import annotations

import io
import json
from urllib.parse import quote

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.data_channel.curated.approved_version_reader import (
    ReviewApprovalError,
    latest_dataset_version,
    load_all_rows_with_edits,
    require_version_approved,
    version_review,
)
from app.data_channel.datasets.models import Dataset
from app.data_channel.datasets.service import DatasetReadError, rows_to_csv_bytes


def require_current_approved_for_read(
    db: Session,
    dataset_id: str,
    *,
    action: str,
):
    """Require the canonical current version to have approval evidence."""
    version = latest_dataset_version(db, dataset_id)
    if version is None:
        raise HTTPException(
            409,
            detail={
                "code": "dataset_version_not_approved",
                "message": f"该成品数据集尚无可用于{action}的数据版本。",
            },
        )
    try:
        require_version_approved(db, dataset_id, version)
    except ReviewApprovalError as exc:
        review = version_review(db, dataset_id, version)
        rejected = review is not None and review.status == "rejected"
        raise HTTPException(
            409,
            detail={
                "code": (
                    "dataset_version_rejected"
                    if rejected
                    else "dataset_version_not_approved"
                ),
                "message": (
                    (
                        f"当前数据版本 v{version.version_no} 已拒绝，"
                        f"仅保留用于审核审计，不能用于普通{action}或进入本体。"
                    )
                    if rejected
                    else (
                        f"当前数据版本 v{version.version_no} 尚未通过审核，"
                        f"不能用于普通{action}或进入本体。"
                    )
                ),
                "dataset_version_id": version.id,
                "version_no": version.version_no,
                "review_status": (
                    review.status if review is not None else "pending_review"
                ),
            },
        ) from exc
    return version


def preview_curated(
    db: Session,
    dataset_id: str,
    *,
    limit: int,
    offset: int,
) -> dict:
    """Return a page from the approved current version."""
    dataset = (
        db.query(Dataset)
        .filter(Dataset.id == dataset_id, Dataset.kind == "curated")
        .first()
    )
    if not dataset:
        raise HTTPException(404, "Curated dataset not found")
    version = require_current_approved_for_read(
        db,
        dataset_id,
        action="预览",
    )

    try:
        from app.data_channel.curated.approved_version_reader import (
            apply_row_edits_to_batch,
        )
        from app.data_channel.datasets import lake_store

        if lake_store.version_uses_lake(dataset, version):
            # 物理表真分页：当前版本（即最新审批版本）就是物理表当前状态；
            # 行编辑按页叠加（写入时已对绑定版本做过存在性校验）
            total_rows = int(version.rowcount or 0)
            rows = lake_store.page_rows(db, dataset, offset, limit)
            rows = apply_row_edits_to_batch(
                db, dataset_id, rows, dataset_version_id=version.id)
        else:
            all_rows = load_all_rows_with_edits(
                db,
                dataset_id,
                require_approved=True,
                version=version,
            )
            total_rows = len(all_rows)
            rows = all_rows[offset : offset + limit]
        schema_columns = (dataset.schema_json or {}).get("columns") or []
        columns = [
            (
                str(item.get("name") or "")
                if isinstance(item, dict)
                else str(item)
            )
            for item in schema_columns
        ]
        columns = [column for column in columns if column]
        if not columns and rows:
            columns = list(rows[0].keys())
        return {
            "dataset_id": dataset_id,
            "name": dataset.name,
            "rows": rows,
            "count": len(rows),
            "columns": columns,
            "total_rows": total_rows,
            "offset": offset,
            "limit": limit,
            "has_more": offset + len(rows) < total_rows,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(502, f"成品数据读取失败：{exc}") from exc


def export_curated(
    db: Session,
    dataset_id: str,
    *,
    output_format: str,
) -> StreamingResponse:
    """Export all rows from the approved current version as CSV or XLSX."""
    dataset = (
        db.query(Dataset)
        .filter(Dataset.id == dataset_id, Dataset.kind == "curated")
        .first()
    )
    if not dataset:
        raise HTTPException(404, "Curated dataset not found")
    version = require_current_approved_for_read(
        db,
        dataset_id,
        action="生产导出",
    )

    from app.data_channel.datasets import lake_store

    if lake_store.version_uses_lake(dataset, version):
        return _export_curated_from_lake(
            db, dataset, version, output_format)

    try:
        rows = load_all_rows_with_edits(
            db,
            dataset_id,
            require_approved=True,
            version=version,
        )
    except DatasetReadError as exc:
        raise HTTPException(502, f"成品数据导出失败：{exc}") from exc
    except ValueError as exc:
        raise HTTPException(
            409,
            detail={
                "code": "review_edit_identity_error",
                "message": str(exc),
            },
        ) from exc

    schema = dataset.schema_json if isinstance(dataset.schema_json, dict) else {}
    columns: list[str] = []

    def add_column(value) -> None:
        name = str(value or "").strip()
        if name and name not in columns:
            columns.append(name)

    for item in schema.get("columns") or []:
        add_column(item.get("name") if isinstance(item, dict) else item)
    for item in schema.get("columns_typed") or []:
        if isinstance(item, dict):
            add_column(item.get("name"))
    for row in rows:
        for name in row:
            add_column(name)

    safe_name = (
        "".join(
            character
            for character in dataset.name
            if character not in '\\/:*?"<>|'
        ).strip()
        or "成品数据集"
    )
    filename = f"{safe_name}.{output_format}"
    disposition = f"attachment; filename*=UTF-8''{quote(filename)}"

    if output_format == "csv":
        data = b"\xef\xbb\xbf" + rows_to_csv_bytes(rows, columns)
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": disposition},
        )

    import openpyxl

    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet(title="数据")
    sheet.append(columns)
    for row in rows:
        values = []
        for column in columns:
            value = row.get(column, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            values.append(value)
        sheet.append(values)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": disposition},
    )


def _export_cell_text(value):
    """与 rows_to_csv_bytes 的单元格口径一致：None→""，嵌套压 JSON。"""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _export_curated_from_lake(
    db: Session,
    dataset,
    version,
    output_format: str,
) -> StreamingResponse:
    """湖表版本的流式导出：stream_rows 分批 + 批内叠加已批准行编辑。

    列集合取契约列（物理表行恒等于契约列集，不需要全量行扫描）：schema
    columns ∪ columns_typed，兜底 lake_columns / 首批行键。CSV 增量产出
    （BOM + 分批写）；xlsx 用 openpyxl write_only + SpooledTemporaryFile。
    """
    from app.data_channel.curated.approved_version_reader import (
        apply_row_edits_to_batch,
    )
    from app.data_channel.datasets import lake_store

    schema = dataset.schema_json if isinstance(dataset.schema_json, dict) else {}
    columns: list[str] = []

    def add_column(value) -> None:
        name = str(value or "").strip()
        if name and name not in columns:
            columns.append(name)

    for item in schema.get("columns") or []:
        add_column(item.get("name") if isinstance(item, dict) else item)
    for item in schema.get("columns_typed") or []:
        if isinstance(item, dict):
            add_column(item.get("name"))
    if not columns:
        for name in (schema.get("lake_columns") or {}):
            add_column(name)

    # 编辑叠加的失败契约（409）必须在响应开始流式输出前暴露：先取首批试叠加
    first_batch: list[dict] = []
    batches = lake_store.stream_rows(db, dataset, batch_size=5000)
    try:
        first_batch = next(batches, None) or []
        if not columns and first_batch:
            for name in first_batch[0]:
                add_column(name)
        first_batch = apply_row_edits_to_batch(
            db, dataset.id, first_batch, dataset_version_id=version.id)
    except ValueError as exc:
        raise HTTPException(
            409,
            detail={
                "code": "review_edit_identity_error",
                "message": str(exc),
            },
        ) from exc
    except Exception as exc:
        raise HTTPException(502, f"成品数据导出失败：{exc}") from exc

    def rest_batches():
        yield first_batch
        for batch in batches:
            yield apply_row_edits_to_batch(
                db, dataset.id, batch, dataset_version_id=version.id)

    safe_name = (
        "".join(
            character
            for character in dataset.name
            if character not in '\\/:*?"<>|'
        ).strip()
        or "成品数据集"
    )
    filename = f"{safe_name}.{output_format}"
    disposition = f"attachment; filename*=UTF-8''{quote(filename)}"

    if output_format == "csv":
        import csv as _csv

        def iter_csv():
            yield b"\xef\xbb\xbf"
            header_done = False
            for batch in rest_batches():
                if not batch:
                    continue
                buf = io.StringIO()
                writer = _csv.DictWriter(
                    buf, fieldnames=columns, extrasaction="ignore", restval="")
                if not header_done:
                    writer.writeheader()
                    header_done = True
                for row in batch:
                    writer.writerow(
                        {c: _export_cell_text(row.get(c)) for c in columns})
                yield buf.getvalue().encode("utf-8")

        return StreamingResponse(
            iter_csv(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": disposition},
        )

    import tempfile

    import openpyxl

    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet(title="数据")
    sheet.append(columns)
    try:
        for batch in rest_batches():
            for row in batch:
                sheet.append([
                    (json.dumps(value, ensure_ascii=False)
                     if isinstance(value := row.get(column, ""), (dict, list))
                     else value)
                    for column in columns
                ])
        # write_only 工作簿流式落临时文件（内存上限外溢到磁盘），再分块回传
        spool = tempfile.SpooledTemporaryFile(max_size=32 * 1024 * 1024)
        workbook.save(spool)
    except Exception as exc:
        raise HTTPException(502, f"成品数据导出失败：{exc}") from exc
    spool.seek(0)
    return StreamingResponse(
        iter(lambda: spool.read(1 << 16), b""),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": disposition},
    )
