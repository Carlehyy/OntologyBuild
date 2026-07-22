"""Dataset CRUD + 版本管理服务 (含 DuckDB 预览)"""
from __future__ import annotations
import hashlib
import json
import logging
import re
import uuid
from contextlib import nullcontext
from datetime import datetime, timezone
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.v2.dataset import (
    Dataset, DatasetVersion, DatasetVersionEvent, StorageDeletionOutbox,
)
from app.services.storage_service import (
    StorageService,
    get_environment_storage_service,
    get_storage_service,
)

logger = logging.getLogger(__name__)


class DatasetReadError(RuntimeError):
    """数据集内容读取/解析失败。

    与「数据集还没有数据」（返回空列表）严格区分：合并基座等场景把读失败
    当空列表处理，会让 append/upsert 把湖中存量折叠成本次增量。
    """


def dataset_kind_uses_database(kind: str | None) -> bool:
    """Whether a dataset version belongs in the platform database.

    Only genuinely unstructured source files stay in object storage.  Tabular,
    semi-structured and curated lake snapshots are transactional database data;
    administrator MinIO configuration must never redirect them.
    """
    return str(kind or "").strip().lower() != "unstructured"


def version_has_content(version: DatasetVersion | None) -> bool:
    """Return payload presence without materializing a deferred database blob."""
    return bool(version and (
        version.data_size is not None or bool(version.storage_uri)
    ))


class TabularParseError(ValueError):
    """上传的表格无法可靠解析。

    调用方可把该异常安全地返回为 4xx；不要在 Excel 解析失败后继续把二进制
    内容当 CSV 猜测，否则真正原因会被 ``csv.Error`` 的换行提示覆盖。
    """


def rows_to_csv_bytes(rows: list[dict], columns: list[str]) -> bytes:
    """行列表 → CSV bytes（人工数据集在线编辑的写回路径）。

    列序由调用方给定（保持编辑前的展示顺序）；None → 空串，嵌套结构压 JSON。
    """
    import csv
    import io

    def _cell(v) -> str:
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return str(v)

    if not rows:
        return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore", restval="")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: _cell(row.get(c)) for c in columns})
    return buf.getvalue().encode("utf-8")


def rows_to_parquet_bytes(rows: list[dict]) -> bytes:
    """行列表 → Parquet bytes：湖内产物快照的存储格式（替代 CSV）。

    全列以「字符串」落盘，精确复刻既有 CSV 入湖的往返语义，保证只换存储格式、
    不改变任何下游可见行为：
      - None → ""；bytes → "<N bytes>"；dict/list → 紧凑 JSON；其余 → str(v)
      - 跳过二进制 content 列（与 CSV 路径一致）
      - 列序按首现；行缺某列 → ""（对齐 CSV DictWriter 的 restval）
    压缩用 zstd（结构化数据通常比 CSV 小数倍，直接省存储与读写 I/O）。
    空行 / 无有效列 → b""（读回为 []，与 _safe_csv_bytes 一致）。
    """
    if not rows:
        return b""
    import io as _io
    import pyarrow as pa
    import pyarrow.parquet as pq

    def _cell(v) -> str:
        if v is None:
            return ""
        if isinstance(v, (bytes, bytearray)):
            return f"<{len(v)} bytes>"
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return str(v)

    all_keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row.keys():
            if k == "content" or k in seen:
                continue
            all_keys.append(k)
            seen.add(k)
    if not all_keys:
        return b""

    table = pa.table({
        k: pa.array([_cell(row.get(k)) for row in rows], type=pa.string())
        for k in all_keys
    })
    buf = _io.BytesIO()
    pq.write_table(table, buf, compression="zstd")
    return buf.getvalue()


def _parse_parquet_rows(raw: bytes, limit: int | None, offset: int = 0) -> list[dict]:
    """读 Parquet 字节为行列表。值以字符串返回（写入时已按 CSV 往返语义规范化）。

    offset/limit 走 record-batch 级跳过与提前终止：分页只物化窗口内的行，
    把「翻一页要全量解析成 dict」降到近 O(页)。limit=None 且 offset=0 时全量读。
    """
    import io as _io
    import pyarrow.parquet as pq

    src = _io.BytesIO(raw)
    if limit is None and offset == 0:
        return pq.read_table(src).to_pylist()

    out: list[dict] = []
    skip = offset
    for batch in pq.ParquetFile(src).iter_batches():
        n = batch.num_rows
        if skip >= n:
            skip -= n
            continue
        if skip:
            batch = batch.slice(skip)
            skip = 0
        rows = batch.to_pylist()
        if limit is None:
            out.extend(rows)
        else:
            out.extend(rows[: limit - len(out)])
            if len(out) >= limit:
                break
    return out


_OLE_COMPOUND_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_HYPERLINK_FORMULA = re.compile(
    r'^\s*=?(?:_xlfn\.)?HYPERLINK\(\s*"((?:[^"]|"")*)"\s*[,;]\s*'
    r'"((?:[^"]|"")*)"\s*\)\s*$',
    re.IGNORECASE,
)


def _formula_fallback(value) -> str:
    """为没有缓存计算结果的公式提供不丢数据的文本回退。

    openpyxl 不负责计算公式。Excel 正常保存的工作簿通常带缓存值；若缓存缺失，
    HYPERLINK 优先提取用户可见的显示文本，其他公式保留表达式供后续排查。
    """
    formula = "" if value is None else str(value)
    match = _HYPERLINK_FORMULA.fullmatch(formula)
    if match:
        url, label = (part.replace('""', '"') for part in match.groups())
        return label or url
    return formula


def _row_is_empty(values: list) -> bool:
    return all(value is None or value == "" for value in values)


def _headers_from_row(values: list) -> list[str]:
    headers: list[str] = []
    for index, value in enumerate(values):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        text = "" if value is None else str(value).strip()
        headers.append(text or f"col_{index}")
    return headers


def _collect_sheet_rows(row_iter, *, limit: int | None,
                        offset: int = 0) -> tuple[list[str], list[dict]]:
    """把工作表行流转成表头和记录，并忽略纯空白行。"""
    headers: list[str] = []
    rows: list[dict] = []
    data_index = 0
    for raw_values in row_iter:
        values = list(raw_values)
        if _row_is_empty(values):
            continue
        if not headers:
            headers = _headers_from_row(values)
            if limit is not None and limit <= 0:
                break
            continue
        if data_index < offset:
            data_index += 1
            continue
        data_index += 1
        rows.append({
            header: values[index] if index < len(values) and values[index] is not None else ""
            for index, header in enumerate(headers)
        })
        if limit is not None and len(rows) >= limit:
            break
    return headers, rows


def _xlsx_table(raw: bytes, *, limit: int | None,
                offset: int = 0) -> tuple[list[str], list[dict]]:
    """读取 OOXML Excel；数据值优先，公式无缓存时保留可理解的回退值。"""
    import io
    import openpyxl

    values_wb = None
    formulas_wb = None
    try:
        values_wb = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True, keep_links=False)
        if not values_wb.worksheets:
            raise TabularParseError("Excel 工作簿中没有可读取的工作表")
        values_ws = values_wb.worksheets[0]

        # limit=0 只读取列名，无需为公式回退再打开一次工作簿。
        formulas_ws = None
        if limit != 0:
            formulas_wb = openpyxl.load_workbook(
                io.BytesIO(raw), read_only=True, data_only=False, keep_links=False)
            formulas_ws = formulas_wb.worksheets[0]

        formula_rows = iter(formulas_ws.iter_rows()) if formulas_ws is not None else None

        def _rows():
            for value_row in values_ws.iter_rows():
                formula_row = next(formula_rows, ()) if formula_rows is not None else ()
                width = max(len(value_row), len(formula_row))
                values = []
                for index in range(width):
                    cached = value_row[index].value if index < len(value_row) else None
                    formula_cell = formula_row[index] if index < len(formula_row) else None
                    if (cached is None and formula_cell is not None
                            and formula_cell.data_type == "f"):
                        cached = _formula_fallback(formula_cell.value)
                    values.append(cached)
                yield values

        return _collect_sheet_rows(_rows(), limit=limit, offset=offset)
    except TabularParseError:
        raise
    except Exception as exc:
        detail = str(exc).strip() or type(exc).__name__
        raise TabularParseError(
            "Excel 工作簿解析失败；文件可能已损坏、已加密，或实际格式不是 XLSX"
            f"（{detail}）") from exc
    finally:
        if formulas_wb is not None:
            formulas_wb.close()
        if values_wb is not None:
            values_wb.close()


def _xls_cell_value(book, cell):
    import xlrd

    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, book.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(cell.value, f"#ERROR({cell.value})")
    return cell.value


def _xls_table(raw: bytes, *, limit: int | None,
               offset: int = 0) -> tuple[list[str], list[dict]]:
    """读取 OLE/BIFF 旧版 XLS；xlrd 返回公式的已缓存计算结果。"""
    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=raw, on_demand=True)
        try:
            if workbook.nsheets < 1:
                raise TabularParseError("Excel 工作簿中没有可读取的工作表")
            sheet = workbook.sheet_by_index(0)
            row_iter = (
                [_xls_cell_value(workbook, sheet.cell(row_index, column_index))
                 for column_index in range(sheet.ncols)]
                for row_index in range(sheet.nrows)
            )
            return _collect_sheet_rows(row_iter, limit=limit, offset=offset)
        finally:
            workbook.release_resources()
    except TabularParseError:
        raise
    except Exception as exc:
        detail = str(exc).strip() or type(exc).__name__
        raise TabularParseError(
            "旧版 Excel 工作簿解析失败；文件可能已损坏、已加密，或实际格式不是 XLS"
            f"（{detail}）") from exc


def _decode_table_text(raw: bytes) -> str:
    """解码常见 Excel/中文环境导出的 CSV，拒绝静默替换坏字节。"""
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig")
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        return raw.decode("utf-16")
    for encoding in ("utf-8", "gb18030"):
        try:
            text = raw.decode(encoding)
            if "\x00" in text:
                continue
            return text
        except UnicodeDecodeError:
            continue
    raise TabularParseError(
        "CSV 文本编码无法识别；请另存为 UTF-8 CSV，或直接上传 XLSX/XLS 文件")


def _csv_table(text: str, *, limit: int | None,
               offset: int = 0) -> tuple[list[str], list[dict]]:
    """读取 CSV/TSV；newline='' 同时兼容 CRLF、LF 与旧式 CR 换行。"""
    import csv
    import io

    stream = io.StringIO(text, newline="")
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|") if sample else csv.excel
    except csv.Error:
        dialect = csv.excel
    try:
        reader = csv.reader(stream, dialect=dialect)
        return _collect_sheet_rows(reader, limit=limit, offset=offset)
    except csv.Error as exc:
        raise TabularParseError(
            f"CSV 解析失败（第 {reader.line_num} 行附近）：{exc}") from exc


def _json_rows(text: str, *, limit: int | None, offset: int = 0) -> list[dict]:
    data = json.loads(text)
    if isinstance(data, list):
        return data[offset:] if limit is None else data[offset:offset + limit]
    if isinstance(data, dict):
        return [data] if offset == 0 and (limit is None or limit > 0) else []
    return []


def _parse_stored_rows(raw: bytes, limit: int | None, offset: int = 0) -> list[dict]:
    """把存储对象解析成行列表（Parquet / JSON / Excel / CSV 自动检测）。

    limit=None 表示不设行数上限——合并基座必须全量，截断即丢数据。
    格式按魔数/内容嗅探：新版产物为 Parquet，存量历史 CSV/JSON/xlsx 一并可读，
    支持滚动窗口内的多格式共存（湖版本随流水线运行逐步自然迁移到 Parquet）。
    """
    offset = max(0, offset)
    if not raw:
        return []

    # Parquet（列存，魔数 PAR1）：必须在 UTF-8 解码之前判定，避免把大二进制
    # 整份 decode 成文本。
    if raw[:4] == b"PAR1":
        return _parse_parquet_rows(raw, limit, offset)

    # OOXML（XLSX）和 OLE/BIFF（XLS）必须先于文本解码识别。解析失败时直接
    # 报 Excel 的原始语义，绝不再把二进制内容回退为 CSV。
    if raw[:2] == b"PK":
        return _xlsx_table(raw, limit=limit, offset=offset)[1]
    if raw.startswith(_OLE_COMPOUND_SIGNATURE):
        return _xls_table(raw, limit=limit, offset=offset)[1]

    text = _decode_table_text(raw)
    stripped = text.lstrip()
    if stripped.startswith("[") or stripped.startswith("{"):
        return _json_rows(text, limit=limit, offset=offset)
    return _csv_table(text, limit=limit, offset=offset)[1]


def stored_columns(raw: bytes) -> list[str]:
    """读取对象的物理列，即使对象只有表头、没有数据行也能返回。

    主键契约校验不能把“零行”误当成“列一定存在”；否则仅表头文件可在缺少
    主键列时通过校验，等后续写入才暴露身份断裂。
    """
    if raw[:4] == b"PAR1":
        import io as _io
        import pyarrow.parquet as pq
        return [str(name) for name in pq.read_schema(_io.BytesIO(raw)).names]
    if raw[:2] == b"PK":
        return _xlsx_table(raw, limit=0)[0]
    if raw.startswith(_OLE_COMPOUND_SIGNATURE):
        return _xls_table(raw, limit=0)[0]
    if not raw:
        return []
    text = _decode_table_text(raw)
    stripped = text.lstrip()
    if stripped.startswith("[") or stripped.startswith("{"):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return [str(k) for k in parsed.keys()]
            if isinstance(parsed, list):
                first = next((row for row in parsed if isinstance(row, dict)), None)
                return [str(k) for k in first.keys()] if first else []
        except json.JSONDecodeError as exc:
            raise TabularParseError(f"JSON 解析失败：{exc}") from exc
    return _csv_table(text, limit=0)[0]


def enqueue_storage_deletions(db: Session, storage_uris: list[str]) -> list[str]:
    """把对象清理记录加入调用方当前事务，不在这里提交。

    同一次资产删除里的重复 URI 会合并；跨事务重复记录允许存在，因为删除对象是
    幂等操作。刻意不使用 URI 唯一约束，避免两个资产历史上误共享同一 URI 时，
    outbox 唯一键竞争反过来让元数据删除事务失败。
    """
    normalized = sorted({
        str(uri).strip() for uri in storage_uris
        if uri is not None and str(uri).strip()
    })
    for uri in normalized:
        db.add(StorageDeletionOutbox(storage_uri=uri))
    if normalized:
        db.flush()
    return normalized


def enqueue_dataset_storage_deletions(db: Session, dataset_id: str) -> list[str]:
    """收集 DatasetVersion、媒体原件和 OCR 结果 URI 并写入同事务 outbox。"""
    from app.models.v2.dataset import MediaItem
    from app.data_channel.file_assets.models import PipelineFileAsset

    versions = db.query(DatasetVersion.id, DatasetVersion.storage_uri).filter(
        DatasetVersion.dataset_id == dataset_id).all()
    version_ids = [row.id for row in versions]
    uris = [row.storage_uri for row in versions if row.storage_uri]
    if version_ids:
        media = db.query(MediaItem.storage_uri, MediaItem.ocr_result_uri).filter(
            MediaItem.dataset_version_id.in_(version_ids)).all()
        for row in media:
            if row.storage_uri:
                uris.append(row.storage_uri)
            if row.ocr_result_uri:
                uris.append(row.ocr_result_uri)
        pipeline_files = db.query(PipelineFileAsset.storage_uri).filter(
            PipelineFileAsset.dataset_version_id.in_(version_ids),
            PipelineFileAsset.storage_uri.is_not(None),
        ).all()
        uris.extend(row.storage_uri for row in pipeline_files if row.storage_uri)
    return enqueue_storage_deletions(db, uris)


def _storage_uri_is_referenced(db: Session, storage_uri: str) -> bool:
    """共享 URI 防护：只要还有任一资产元数据引用，就暂不物理删除。"""
    from app.models.v2.dataset import MediaItem
    from app.data_channel.file_assets.models import PipelineFileAsset

    if db.query(DatasetVersion.id).filter(
            DatasetVersion.storage_uri == storage_uri).first() is not None:
        return True
    if db.query(MediaItem.id).filter(
        (MediaItem.storage_uri == storage_uri)
        | (MediaItem.ocr_result_uri == storage_uri)
    ).first() is not None:
        return True
    return db.query(PipelineFileAsset.id).filter(
        PipelineFileAsset.storage_uri == storage_uri,
        PipelineFileAsset.status.in_(("ready", "committed")),
    ).first() is not None


def _deletion_storage_candidates(
    storage_uri: str, storage: StorageService | None,
) -> list[StorageService]:
    """Resolve the store(s) that may own an outbox object.

    File assets are always managed-MinIO objects.  A historical dataset object
    can predate or postdate the configurable-MinIO regression and therefore may
    exist in either endpoint; delete it from both before acknowledging the task.
    An explicit storage override preserves deterministic test/caller behavior.
    """
    if storage is not None:
        candidates = [storage]
    elif storage_uri.startswith("s3://raw-datasets/datasets/"):
        candidates = [
            get_environment_storage_service(),
            get_storage_service(),
        ]
    else:
        candidates = [get_storage_service()]

    unique: list[StorageService] = []
    seen: set[int] = set()
    for candidate in candidates:
        if id(candidate) in seen:
            continue
        seen.add(id(candidate))
        unique.append(candidate)
    return unique


def drain_storage_deletion_outbox(
    db: Session,
    storage: StorageService | None = None,
    *,
    limit: int = 100,
    strict_schema: bool = False,
) -> dict[str, int]:
    """机会式清理 outbox；任何单项失败都保留任务供后续重试。

    对象删除成功后才删除 outbox 行。若此后的数据库提交失败，任务仍会回滚保留，
    下次重试再次删除同一对象也安全。该函数是 best-effort 边界，不把存储故障传播
    成已经提交的 Dataset 删除请求 5xx。
    """
    result = {"deleted": 0, "failed": 0, "deferred": 0}
    try:
        entries = db.query(StorageDeletionOutbox).order_by(
            StorageDeletionOutbox.created_at, StorageDeletionOutbox.id
        ).limit(max(1, min(int(limit), 1000))).all()
    except Exception:
        db.rollback()
        if strict_schema:
            raise
        logger.exception("读取对象存储删除 outbox 失败")
        result["failed"] += 1
        return result
    if not entries:
        return result
    for entry in entries:
        storage_uri = entry.storage_uri
        try:
            if _storage_uri_is_referenced(db, storage_uri):
                entry.last_error = "对象仍被资产元数据引用，已延后物理删除"
                entry.updated_at = datetime.now(timezone.utc)
                db.commit()
                result["deferred"] += 1
                continue

            failures: list[str] = []
            for candidate in _deletion_storage_candidates(storage_uri, storage):
                try:
                    candidate.delete_object(storage_uri)
                except Exception as exc:
                    failures.append(f"{type(exc).__name__}: {exc}")
            if failures:
                entry.attempts = int(entry.attempts or 0) + 1
                entry.last_error = "; ".join(failures)[:2000]
                entry.updated_at = datetime.now(timezone.utc)
                db.commit()
                result["failed"] += 1
                logger.warning(
                    "对象存储删除失败，已保留 outbox 任务（attempt=%s, uri=%s）",
                    entry.attempts, storage_uri)
                continue

            db.delete(entry)
            db.commit()
            result["deleted"] += 1
        except Exception:
            # 数据库提交失败时回滚；即使对象已经被删，outbox 任务仍在并可幂等重试。
            db.rollback()
            result["failed"] += 1
            logger.exception("更新对象存储删除 outbox 状态失败（uri=%s）", storage_uri)
    return result


class DatasetService:
    def __init__(
        self,
        db: Session,
        storage: StorageService | None = None,
        *,
        legacy_storages: list[StorageService] | None = None,
    ):
        self._db = db
        # ``storage`` remains the injection point used by file-backed datasets
        # and tests.  It is intentionally lazy in production so database-backed
        # dataset reads/writes do not depend on MinIO availability.
        self._storage_override = storage
        self._legacy_storages_override = legacy_storages

    def _object_storage(self) -> StorageService:
        """Managed MinIO used only for genuine file payloads."""
        return self._storage_override or get_storage_service()

    def _legacy_storage_candidates(self) -> list[StorageService]:
        """Stores that may contain a pre-database DatasetVersion.

        Versions created before configurable MinIO used the deployment endpoint;
        versions created during the regression may live in managed MinIO.  Try
        both, in that order, and de-duplicate the common no-managed-config case.
        """
        if self._legacy_storages_override is not None:
            candidates = list(self._legacy_storages_override)
        elif self._storage_override is not None:
            candidates = [self._storage_override]
        else:
            candidates = [
                get_environment_storage_service(),
                get_storage_service(),
            ]
        unique: list[StorageService] = []
        seen: set[int] = set()
        for candidate in candidates:
            if id(candidate) in seen:
                continue
            seen.add(id(candidate))
            unique.append(candidate)
        return unique

    def create_dataset(self, name: str, kind: str, connection_id: str | None = None,
                       schema_json: dict | None = None, *,
                       source_resource: str | None = None,
                       producer_pipeline_id: str | None = None,
                       output_key: str | None = None,
                       commit: bool = True) -> Dataset:
        """创建数据集。

        ``commit=False`` 仅供“数据集 + 首版本”原子创建路径使用：先 flush 取得
        UUID，随后由 ``create_version`` 把数据集、版本、latest_version_id 和
        schema 一次提交。默认行为保持兼容。
        """
        ds = Dataset(name=name, kind=kind, source_connection_id=connection_id,
                     source_resource=source_resource,
                     schema_json=schema_json,
                     producer_pipeline_id=producer_pipeline_id,
                     output_key=output_key)
        self._db.add(ds)
        if commit:
            self._db.commit()
            self._db.refresh(ds)
        else:
            self._db.flush()
        return ds

    def create_version(self, dataset_id: str, data: bytes, rowcount: int | None = None,
                       *, schema_json: dict | None = None,
                       _lock_held: bool = False) -> DatasetVersion:
        """原子发布 DatasetVersion。

        结构化、半结构化和成品数据把完整版本载荷与版本元数据放在同一个数据库
        事务中；只有非结构化文件继续写入管理员配置的 MinIO。这样文件存储配置
        变化不会再改变数据流水线与资产湖之间的持久化位置。

        ``_lock_held`` 仅供已经覆盖完整读改写临界区的调用方使用；普通调用一律
        通过数据库锁串行化版本号分配。
        """
        if not _lock_held:
            from app.data_channel.datasets.lock import dataset_write_lock
            guard = dataset_write_lock(
                f"dataset::{dataset_id}", bind=self._db.get_bind(), wait_timeout=300)
        else:
            guard = nullcontext()

        with guard:
            return self._create_version_locked(
                dataset_id, data, rowcount=rowcount, schema_json=schema_json)

    def _create_version_locked(self, dataset_id: str, data: bytes,
                               rowcount: int | None = None,
                               schema_json: dict | None = None) -> DatasetVersion:
        ds = self._db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not ds:
            raise ValueError(f"Dataset {dataset_id} not found")

        # 新版本存完整 SHA-256；load_all_rows 同时兼容历史 16 位校验和。
        checksum = hashlib.sha256(data).hexdigest()
        database_backed = dataset_kind_uses_database(ds.kind)
        ver: DatasetVersion | None = None
        # 版本号 = 查最大值+1，并发写会撞唯一约束 (dataset_id, version_no)——
        # 撞了就重算重试，而不是静默产生重复版本号
        for attempt in range(3):
            last_ver = self._db.query(DatasetVersion).filter(
                DatasetVersion.dataset_id == dataset_id
            ).order_by(DatasetVersion.version_no.desc()).first()
            version_no = (last_ver.version_no + 1) if last_ver else 1

            version_id = str(uuid.uuid4())
            uri: str | None = None
            if not database_backed:
                # 非结构化文件仍使用不可变 UUID key；失败重试不会覆盖已提交版本。
                key = f"datasets/{dataset_id}/objects/{version_id}.bin"
                try:
                    uri = self._object_storage().put_bytes(
                        "raw-datasets", key, data)
                except Exception:
                    # 首版本允许 Dataset 与 Version 同事务创建；对象写失败时必须
                    # 回滚尚未提交的 Dataset，不能在湖里留下无版本空壳。
                    self._db.rollback()
                    raise

            ver = DatasetVersion(
                id=version_id,
                dataset_id=dataset_id,
                version_no=version_no,
                rowcount=rowcount,
                data_blob=data if database_backed else None,
                data_size=len(data) if database_backed else None,
                storage_uri=uri,
                checksum=checksum,
            )
            self._db.add(ver)
            try:
                self._db.flush()
                ds.latest_version_id = ver.id
                # 版本内容与解释它的逻辑契约必须同一事务发布。否则消费者可能
                # 看到新 version，却仍按旧 PK/类型/producer 元数据解释它。
                if schema_json is not None:
                    ds.schema_json = schema_json
                # DatasetVersion 与发布事件同事务提交。后续自动化即使在服务重启、
                # Celery/网络抖动或多实例竞争下失败，也能从 outbox 可靠重试。
                self._db.add(DatasetVersionEvent(
                    dataset_id=dataset_id,
                    dataset_version_id=ver.id,
                    event_type="version_published",
                ))
                self._db.commit()
                break
            except IntegrityError:
                self._db.rollback()
                if uri:
                    try:
                        self._object_storage().delete_object(uri)
                    except Exception:
                        logger.warning(
                            "清理未提交版本对象失败: %s", uri, exc_info=True)
                if attempt == 2:
                    raise
            except Exception:
                self._db.rollback()
                if uri:
                    try:
                        self._object_storage().delete_object(uri)
                    except Exception:
                        logger.warning(
                            "清理未提交版本对象失败: %s", uri, exc_info=True)
                raise
        self._db.refresh(ver)
        self._prune_versions_best_effort(dataset_id)
        return ver

    @staticmethod
    def _checksum_matches(raw: bytes, expected: str | None) -> bool:
        """校验对象内容；兼容两种历史 16 位写法。

        历史 DatasetService 保存全文 SHA-256 的前 16 位，旧 SyncEngine 则只哈希
        前 1KiB。新写入统一保存完整 64 位 SHA-256。
        """
        if not expected:
            return True  # 迁移前允许 checksum=NULL 的存量版本继续读取
        expected = str(expected).lower()
        full = hashlib.sha256(raw).hexdigest()
        if len(expected) == 64:
            return full == expected
        if len(expected) == 16:
            legacy_prefix = hashlib.sha256(raw[:1024]).hexdigest()[:16]
            return full.startswith(expected) or legacy_prefix == expected
        return False

    def _prune_versions_best_effort(self, dataset_id: str) -> None:
        """机会式版本保留：只留最近 N 个全量快照，防止存储无限膨胀。

        任何失败都不影响本次写入；删不掉的对象保留记录，下次写入再试。
        """
        try:
            self._prune_versions(dataset_id)
        except Exception:
            self._db.rollback()
            logger.warning(f"数据集 {dataset_id} 旧版本清理失败（不影响本次写入）", exc_info=True)

    def _prune_versions(self, dataset_id: str) -> None:
        from app.config import settings
        from app.data_channel.file_assets.models import PipelineFileAsset
        from app.models.v2.dataset import MediaItem
        from app.models.v2.curated import CuratedReview

        keep = int(getattr(settings, "dataset_version_keep", 0) or 0)
        if keep <= 0:
            return  # 0/负数 = 不清理
        versions = self._db.query(DatasetVersion).filter(
            DatasetVersion.dataset_id == dataset_id
        ).order_by(DatasetVersion.version_no.desc()).all()
        candidates = versions[keep:]
        if not candidates:
            return
        # 挂着媒体文件（OCR 等）的版本不清理，避免级联删除媒体记录
        media_ver_ids = {row[0] for row in self._db.query(MediaItem.dataset_version_id).filter(
            MediaItem.dataset_version_id.in_([v.id for v in candidates])).all()}
        # 审核背书绑定不可变 DatasetVersion；先删对象再撞 FK 会留下“审核记录还在、
        # 证据快照已丢”的不可恢复状态，因此被任何审核引用的版本永久排除清理。
        reviewed_ver_ids = {row[0] for row in self._db.query(CuratedReview.dataset_version_id).filter(
            CuratedReview.dataset_version_id.in_([v.id for v in candidates])).all()}

        removed = 0
        for v in candidates:
            if v.id in media_ver_ids or v.id in reviewed_ver_ids:
                continue
            storage_uris = [v.storage_uri] if v.storage_uri else []
            # File rows cascade with DatasetVersion. Queue their private MinIO
            # objects before deleting the version so metadata and object
            # lifecycle remain in the same transaction boundary.
            file_assets = self._db.query(PipelineFileAsset).filter(
                PipelineFileAsset.dataset_version_id == v.id,
            ).all()
            storage_uris.extend(
                asset.storage_uri for asset in file_assets if asset.storage_uri)
            if storage_uris:
                # 元数据删除与 outbox 同事务；对象存储失败不再迫使版本元数据永远
                # 留在湖里，也不会造成无追踪的对象泄漏。
                enqueue_storage_deletions(self._db, storage_uris)
            # Do not depend on database-specific FK enforcement here.  Some
            # SQLite deployments/tests cannot enable ON DELETE CASCADE; an
            # explicit delete also ensures the outbox drain no longer sees a
            # committed reference and can reclaim the object immediately.
            for asset in file_assets:
                self._db.delete(asset)
            self._db.delete(v)
            removed += 1
        if removed:
            self._db.commit()
            drain_storage_deletion_outbox(
                self._db, storage=self._storage_override)
            logger.info(f"数据集 {dataset_id} 版本保留清理：删除 {removed} 个旧版本（保留最近 {keep} 个）")

    def get_dataset(self, dataset_id: str) -> Dataset | None:
        return self._db.query(Dataset).filter(Dataset.id == dataset_id).first()

    def list_datasets(self, kind: str | None = None) -> list[Dataset]:
        q = self._db.query(Dataset)
        if kind:
            q = q.filter(Dataset.kind == kind)
        return q.all()

    def list_versions(self, dataset_id: str) -> list[DatasetVersion]:
        return self._db.query(DatasetVersion).filter(
            DatasetVersion.dataset_id == dataset_id
        ).order_by(DatasetVersion.version_no).all()

    def _resolve_version(self, dataset_id: str, version_no: int | None) -> DatasetVersion | None:
        q = self._db.query(DatasetVersion).filter(
            DatasetVersion.dataset_id == dataset_id)
        if version_no is None:
            return q.order_by(DatasetVersion.version_no.desc()).first()
        return q.filter(DatasetVersion.version_no == version_no).first()

    def load_version_bytes(
        self, dataset_id: str, version_no: int | None = None,
    ) -> bytes | None:
        """Strictly load one immutable version from DB or legacy object storage."""
        ver = self._resolve_version(dataset_id, version_no)
        if ver is None:
            return None

        if ver.data_blob is not None:
            raw = bytes(ver.data_blob)
            if not self._checksum_matches(raw, ver.checksum):
                raise DatasetReadError(
                    f"数据集 {dataset_id} v{ver.version_no} 校验和不匹配"
                    "（平台数据库）：版本内容可能已损坏")
            return raw
        elif ver.storage_uri:
            failures: list[str] = []
            for index, storage in enumerate(
                self._legacy_storage_candidates(), start=1,
            ):
                try:
                    raw = storage.get_object(ver.storage_uri)
                except Exception as exc:
                    failures.append(f"{type(exc).__name__}: {exc}")
                    continue
                if self._checksum_matches(raw, ver.checksum):
                    return raw
                # The same legacy URI can exist in both endpoints.  A stale or
                # overwritten copy must not mask a valid copy in the next one.
                failures.append(f"候选存储 {index} 的对象校验和不匹配")

            details = "; ".join(failures) or "没有可用的历史对象存储"
            raise DatasetReadError(
                f"数据集 {dataset_id} v{ver.version_no} 历史存储对象读取失败"
                f"（{ver.storage_uri}）：{details}")
        else:
            return None

    def load_all_rows(self, dataset_id: str, version_no: int | None = None) -> list[dict]:
        """严格全量读：不设行数上限，读失败硬报错（DatasetReadError）。

        入湖合并基座、需要「必须完整」语义的消费方一律用这个，不要用
        preview()——它容错返回空列表，会把「读失败」伪装成「湖是空的」。
        数据集尚无版本/内容为空 → 返回 []（这是合法状态，不是错误）。
        """
        ver = self._resolve_version(dataset_id, version_no)
        if not version_has_content(ver):
            return []
        raw = self.load_version_bytes(dataset_id, ver.version_no)
        if not raw:
            return []
        try:
            return _parse_stored_rows(raw, limit=None)
        except Exception as e:
            raise DatasetReadError(
                f"数据集 {dataset_id} v{ver.version_no} 内容解析失败：{e}") from e

    def preview(self, dataset_id: str, version_no: int | None,
                limit: int = 100, offset: int = 0) -> list[dict]:
        """CSV/JSON 数据预览。无需 DuckDB, 纯 Python 处理。

        version_no=None → 最新版本。消费方（映射/管道/质量报告）除非明确要
        历史版本，一律应传 None——增量同步产生 v2+ 后钉死 v1 会读到陈旧数据。

        offset：跳过前 N 行数据行（表头不计），配合 limit 做分页预览。
        容错语义：读取/解析失败返回 []，仅适合 UI 展示；「必须完整」的
        场景（如合并基座）用 load_all_rows()。
        """
        ver = self._resolve_version(dataset_id, version_no)
        if not version_has_content(ver):
            return []
        try:
            raw = self.load_version_bytes(dataset_id, ver.version_no)
            if not raw:
                return []
            return _parse_stored_rows(raw, limit=limit, offset=offset)
        except Exception:
            return []
