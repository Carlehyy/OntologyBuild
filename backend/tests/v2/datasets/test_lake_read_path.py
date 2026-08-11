"""curated 读路径的物理湖表分流测试。

覆盖：service 层 load_all_rows/preview 分流（最新版真分页/历史版回放/blob
遗留）、混合态 review-diff（current 湖表/previous blob）、preview_curated
真分页、export_curated 流式写出与旧路径字节一致、行编辑按批叠加与全量叠加
对拍、prune 的回放链保留语义、删除级联物理表与变更集。
"""
from __future__ import annotations

import io
import json
import uuid

import pytest
import sqlalchemy as sa

from app.data_channel.curated.approved_version_reader import (
    apply_all_row_edits,
    iter_rows_with_edits,
)
from app.data_channel.curated.review_service import ReviewService
from app.data_channel.datasets import lake_store
from app.data_channel.datasets.lake_gate import split_pk
from app.data_channel.datasets.models import (
    Dataset,
    DatasetChangeset,
    DatasetChangesetRow,
    DatasetVersion,
)
from app.data_channel.datasets.service import (
    DatasetService,
    rows_to_csv_bytes,
    rows_to_parquet_bytes,
    version_has_content,
)
from app.main import app
from app.models.v2.curated import CuratedReview
from app.routers.v2 import curated as curated_module
from app.routers.v2 import datasets as datasets_module
from app.routers.v2 import mappings as mappings_module


@pytest.fixture(autouse=True)
def _drop_lake_tables(db):
    """物理湖表不在 Base.metadata 中（运行时 DDL），逐测试清理避免串库。"""
    yield
    conn = db.connection()
    for name in sa.inspect(conn).get_table_names():
        if name.startswith("lake_ds_"):
            conn.execute(sa.text(f'DROP TABLE IF EXISTS "{name}"'))
    db.commit()


@pytest.fixture
def api(client, db):
    def _override():
        yield db
    for mod in (datasets_module, mappings_module, curated_module):
        app.dependency_overrides[mod.get_db] = _override
    yield client
    for mod in (datasets_module, mappings_module, curated_module):
        app.dependency_overrides.pop(mod.get_db, None)


def _make_lake_dataset(db, runs, pk="id", columns=("id", "name")):
    """按序多次 upsert_run 建湖表资产，runs: list[(write_mode, rows)]。"""
    ds = Dataset(
        id=str(uuid.uuid4()), name=f"读路径-{uuid.uuid4().hex[:6]}",
        kind="curated",
        schema_json={"primary_key": pk, "columns": list(columns)})
    db.add(ds)
    db.commit()
    versions = []
    for mode, rows in runs:
        version, _ = lake_store.upsert_run(db, ds, rows, mode, split_pk(pk))
        versions.append(version)
    db.refresh(ds)
    return ds, versions


def _approve_latest(db, dataset_id: str):
    return ReviewService(db).approve(ReviewService(db).start_review(dataset_id).id)


# ── service 层分流 ───────────────────────────────────────────
def test_load_all_rows_and_preview_reroute_to_lake(db):
    ds, (v1, v2) = _make_lake_dataset(db, [
        ("overwrite", [{"id": "1", "name": "a"}, {"id": "2", "name": "b"},
                       {"id": "3", "name": "c"}]),
        ("upsert", [{"id": "2", "name": "B"}, {"id": "4", "name": "d"}]),
    ])
    svc = DatasetService(db)

    assert version_has_content(v2) is True
    latest_rows = svc.load_all_rows(ds.id)
    assert {r["id"] for r in latest_rows} == {"1", "2", "3", "4"}
    assert next(r for r in latest_rows if r["id"] == "2")["name"] == "B"
    # 历史湖表版本经变更集回放
    v1_rows = svc.load_all_rows(ds.id, v1.version_no)
    assert {r["id"] for r in v1_rows} == {"1", "2", "3"}
    assert next(r for r in v1_rows if r["id"] == "2")["name"] == "b"

    # 最新版真分页（主键序）；历史版回放后切片
    assert [r["id"] for r in svc.preview(ds.id, None, limit=2, offset=1)] == ["2", "3"]
    assert [r["id"] for r in svc.preview(ds.id, v1.version_no, limit=2)] == ["1", "2"]


def test_load_all_rows_blob_version_stays_legacy(db):
    ds, _ = _make_lake_dataset(db, [])
    svc = DatasetService(db)
    svc.create_version(ds.id, rows_to_parquet_bytes(
        [{"id": "1", "name": "blob"}]), rowcount=1)
    assert svc.load_all_rows(ds.id) == [{"id": "1", "name": "blob"}]
    assert svc.preview(ds.id, None, limit=10) == [{"id": "1", "name": "blob"}]


def test_version_has_content_semantics(db):
    assert version_has_content(None) is False
    # 非 curated 的 blob 版本（data_size 非空）不受影响
    legacy = DatasetVersion(
        id=str(uuid.uuid4()), dataset_id="x", version_no=1,
        rowcount=1, data_blob=b"[]", data_size=2, checksum="ck")
    assert version_has_content(legacy) is True
    # 完全无载荷也无行数的版本：无内容（既有语义）
    empty = DatasetVersion(
        id=str(uuid.uuid4()), dataset_id="x", version_no=2)
    assert version_has_content(empty) is False


# ── 混合态 review-diff（current 湖表 / previous blob）─────────
def test_review_diff_mixed_state_current_lake_previous_blob(api, auth_headers, db):
    svc = DatasetService(db)
    ds, _ = _make_lake_dataset(db, [])
    svc.create_version(ds.id, rows_to_parquet_bytes(
        [{"id": "1", "name": "a"}, {"id": "2", "name": "b"}]), rowcount=2)
    # v2 走湖表写入（先经懒引导灌入遗留基座）：1 删、2 改、3 增
    svc.bootstrap_lake_base(ds)
    v2, _ = lake_store.upsert_run(db, ds, [
        {"id": "2", "name": "B"}, {"id": "3", "name": "c"},
    ], "overwrite", ["id"])

    r = api.get(f"/api/v2/curated/{ds.id}/review-diff", headers=auth_headers)
    assert r.status_code == 200, r.text
    body = r.json().get("data", r.json())
    assert body["current"]["version_no"] == 2
    assert body["previous"]["version_no"] == 1
    assert body["current"]["total"] == 2
    assert body["previous"]["total"] == 2
    assert {r["id"] for r in body["current"]["rows"]} == {"2", "3"}
    assert {r["id"] for r in body["previous"]["rows"]} == {"1", "2"}

    delta = body["delta"]
    assert (delta["added_count"], delta["updated_count"],
            delta["deleted_count"]) == (1, 1, 1)
    assert delta["added_sample"][0]["id"] == "3"
    assert delta["deleted_sample"][0]["id"] == "1"
    assert delta["deleted_sample"][0]["name"] == "a"
    assert delta["updated_sample"][0]["before"]["name"] == "b"
    assert delta["updated_sample"][0]["after"]["name"] == "B"
    assert body["row_pk_encoding"] == "plain-string"
    assert body["current_row_pks"] == ["2", "3"]


# ── preview/export 真分页与流式 ──────────────────────────────
def test_preview_curated_true_pagination_on_lake(api, auth_headers, db):
    ds, _ = _make_lake_dataset(db, [
        ("overwrite", [{"id": str(i), "name": f"n{i}"} for i in range(5)]),
    ])
    _approve_latest(db, ds.id)

    pages = []
    for offset in (0, 2, 4):
        r = api.get(f"/api/v2/curated/{ds.id}/preview?limit=2&offset={offset}",
                    headers=auth_headers)
        assert r.status_code == 200, r.text
        pages.append(r.json().get("data", r.json()))
    assert [p["count"] for p in pages] == [2, 2, 1]
    assert [p["has_more"] for p in pages] == [True, True, False]
    assert all(p["total_rows"] == 5 for p in pages)
    assert [r["id"] for p in pages for r in p["rows"]] == ["0", "1", "2", "3", "4"]
    assert pages[0]["columns"] == ["id", "name"]


def test_export_curated_csv_and_xlsx_from_lake(api, auth_headers, db):
    ds, _ = _make_lake_dataset(db, [
        ("overwrite", [{"id": "1", "name": "审核前"},
                       {"id": "2", "name": "保持不变"}]),
    ])
    svc = ReviewService(db)
    review = svc.start_review(ds.id)
    svc.batch_edit_rows(review.id, [{
        "row_pk": "1", "field_name": "name",
        "old_value": "审核前", "new_value": "审核后",
    }])
    svc.approve(review.id)

    csv_response = api.get(
        f"/api/v2/curated/{ds.id}/export?format=csv", headers=auth_headers)
    assert csv_response.status_code == 200, csv_response.text
    expected_rows = [{"id": "1", "name": "审核后"}, {"id": "2", "name": "保持不变"}]
    assert csv_response.content == (
        b"\xef\xbb\xbf" + rows_to_csv_bytes(expected_rows, ["id", "name"]))

    xlsx_response = api.get(
        f"/api/v2/curated/{ds.id}/export?format=xlsx", headers=auth_headers)
    assert xlsx_response.status_code == 200, xlsx_response.text
    import openpyxl
    workbook = openpyxl.load_workbook(
        io.BytesIO(xlsx_response.content), read_only=True)
    assert list(workbook.active.iter_rows(values_only=True)) == [
        ("id", "name"), ("1", "审核后"), ("2", "保持不变")]


# ── 行编辑按批叠加对拍 ──────────────────────────────────────
def test_batch_overlay_matches_full_overlay(db):
    rows = [{"id": str(i), "name": f"n{i}"} for i in range(10)]
    ds, _ = _make_lake_dataset(db, [("overwrite", rows)])
    svc = ReviewService(db)
    review = svc.start_review(ds.id)
    svc.batch_edit_rows(review.id, [
        {"row_pk": "1", "field_name": "name", "old_value": "n1", "new_value": "甲"},
        {"row_pk": "9", "field_name": "name", "old_value": "n9", "new_value": "乙"},
    ])
    svc.approve(review.id)

    lake_rows = [row for batch in lake_store.stream_rows(db, ds, batch_size=3)
                 for row in batch]
    expected = apply_all_row_edits(db, ds.id, lake_rows, dataset_version_id=None)
    streamed = [row for batch in iter_rows_with_edits(db, ds.id, batch_size=3)
                for row in batch]
    assert streamed == expected
    assert streamed[1]["name"] == "甲" and streamed[9]["name"] == "乙"


# ── prune：回放链感知保留 ────────────────────────────────────
def _run_n_versions(db, ds, n: int, start: int = 0):
    versions = []
    for i in range(start, start + n):
        version, _ = lake_store.upsert_run(
            db, ds, [{"id": "1", "name": f"v{i + 1}"}], "overwrite", ["id"])
        versions.append(version)
    return versions


def test_prune_deletes_old_versions_and_their_changesets(db, monkeypatch):
    from app.config import settings
    monkeypatch.setattr(settings, "dataset_version_keep", 3)
    ds, _ = _make_lake_dataset(db, [])
    versions = _run_n_versions(db, ds, 5)
    # 生产上 prune 由调用方在每次发布后机会式执行（pipeline_run 经
    # svc._prune_versions_best_effort）；测试统一触发一次等价验证
    DatasetService(db)._prune_versions_best_effort(ds.id)

    remaining = db.query(DatasetVersion).filter(
        DatasetVersion.dataset_id == ds.id).order_by(DatasetVersion.version_no).all()
    assert [v.version_no for v in remaining] == [3, 4, 5]
    assert db.query(DatasetChangeset).filter(
        DatasetChangeset.version_id.in_([versions[0].id, versions[1].id])).count() == 0
    assert db.query(DatasetChangesetRow).filter(
        DatasetChangesetRow.changeset_id.in_(
            sa.select(DatasetChangeset.id).where(
                DatasetChangeset.dataset_id == ds.id))).count() >= 0
    assert lake_store.count_rows(db, ds) == 1


def test_prune_keeps_replay_chain_for_review_pinned_version(db, monkeypatch):
    """被审核钉住的最旧版本使整条后续变更集链成为回放必需：元数据与变更集
    全部保留，且 rows_at_version 仍能正确回放到被钉版本。"""
    from app.config import settings
    monkeypatch.setattr(settings, "dataset_version_keep", 3)
    ds, _ = _make_lake_dataset(db, [])
    pinned = _run_n_versions(db, ds, 1)[0]
    # 先发版后立即钉住 v1，再继续发版触发 prune
    db.add(CuratedReview(
        id=str(uuid.uuid4()), curated_dataset_id=ds.id,
        dataset_version_id=pinned.id, status="approved"))
    db.commit()
    _run_n_versions(db, ds, 5, start=1)  # v2..v6
    DatasetService(db)._prune_versions_best_effort(ds.id)

    remaining_nos = [v.version_no for v in db.query(DatasetVersion).filter(
        DatasetVersion.dataset_id == ds.id).order_by(DatasetVersion.version_no)]
    assert remaining_nos == [1, 2, 3, 4, 5, 6]  # 钉住 v1 ⇒ 链上全部保留
    assert db.query(DatasetChangeset).filter(
        DatasetChangeset.dataset_id == ds.id).count() == 6
    replayed = lake_store.rows_at_version(db, ds, pinned.version_no)
    assert replayed == [{"id": "1", "name": "v1"}]


# ── 删除级联 ────────────────────────────────────────────────
def test_delete_curated_drops_lake_table_and_changesets(db):
    from app.data_channel.curated.lifecycle_service import delete_curated

    ds, _ = _make_lake_dataset(db, [
        ("overwrite", [{"id": "1", "name": "a"}]),
        ("upsert", [{"id": "2", "name": "b"}]),
    ])
    table_name = lake_store.lake_table_name(ds.id)
    assert lake_store.lake_table_exists(db, ds.id)

    delete_curated(db, ds.id, force=False)

    assert not lake_store.lake_table_exists(db, ds.id)
    assert db.query(DatasetChangesetRow).filter(
        DatasetChangesetRow.changeset_id.in_(
            sa.select(DatasetChangeset.id).where(
                DatasetChangeset.dataset_id == ds.id))).count() == 0
    assert db.query(DatasetChangeset).filter(
        DatasetChangeset.dataset_id == ds.id).count() == 0
    assert db.query(DatasetVersion).filter(
        DatasetVersion.dataset_id == ds.id).count() == 0
