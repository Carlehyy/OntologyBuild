"""在线建表（人工数据集）：不上传文件，在线定义列/类型/主键后逐行维护。

产品定位：与上传创建的人工数据集完全同权——声明主键后可被本体映射灌入、
可作流水线数据源、可上传文件批量补数。区别仅在来源（source=manual）与
列类型语义（用户声明 types_source=declared，编辑时校验且不随数据重推断）。
"""
from __future__ import annotations

import io
import json

import pytest

from app.main import app
from app.routers.v2 import datasets as datasets_module
from app.routers.v2 import mappings as mappings_module


class FakeStorage:
    def __init__(self):
        self.objects: dict[str, bytes] = {}

    def put_bytes(self, bucket, key, data, content_type=""):
        uri = f"s3://{bucket}/{key}"
        self.objects[uri] = data
        return uri

    def get_object(self, uri):
        if uri not in self.objects:
            raise FileNotFoundError(uri)
        return self.objects[uri]

    def delete_object(self, uri):
        self.objects.pop(uri, None)


@pytest.fixture
def fake_storage(monkeypatch):
    from app.data_channel.datasets import service as ds_service
    fs = FakeStorage()
    monkeypatch.setattr(ds_service, "get_storage_service", lambda: fs)
    return fs


@pytest.fixture
def api(client, db, fake_storage):
    def _override():
        yield db

    app.dependency_overrides[datasets_module.get_db] = _override
    app.dependency_overrides[mappings_module.get_db] = _override
    yield client
    app.dependency_overrides.pop(datasets_module.get_db, None)
    app.dependency_overrides.pop(mappings_module.get_db, None)


def _create_table(api, headers, name="设备台账", columns=None, primary_key="编号"):
    payload = {
        "name": name,
        "columns": columns if columns is not None else [
            {"name": "编号", "type": "string"},
            {"name": "名称", "type": "string"},
            {"name": "数量", "type": "integer"},
        ],
        "primary_key": primary_key,
    }
    return api.post("/api/v2/datasets/create-table", json=payload, headers=headers)


# ── 创建 ─────────────────────────────────────────────────────
def test_create_table_basic(api, auth_headers):
    r = _create_table(api, auth_headers)
    assert r.status_code == 201, r.text
    data = r.json()["data"]
    assert data["version_no"] == 1
    assert data["columns"] == ["编号", "名称", "数量"]
    assert data["primary_key"] == "编号"
    ds_id = data["id"]

    # 总览：来源=manual、主键已声明、v1 空表
    ov = api.get("/api/v2/datasets/overview", headers=auth_headers).json()
    item = next(i for i in ov["items"] if i["id"] == ds_id)
    assert item["source"] == "manual"
    assert item["primary_key"] == "编号"
    assert item["version_count"] == 1 and item["rowcount"] == 0

    # 预览：0 行但表头来自契约列（编辑器靠它渲染空表）
    pv = api.get(f"/api/v2/datasets/{ds_id}/preview", headers=auth_headers).json()
    assert pv["columns"] == ["编号", "名称", "数量"]
    assert pv["rows"] == [] and pv["total_rows"] == 0

    # schema：返回声明类型而非推断（空表推断只会得到空列表）
    sc = api.get(f"/api/v2/datasets/{ds_id}/schema", headers=auth_headers).json()
    assert {c["name"]: c["type"] for c in sc["columns"]} == {
        "编号": "string", "名称": "string", "数量": "integer"}


def test_manual_version_subscription_requires_versioned_draft_after_publish(
    api, auth_headers, ontology, db,
):
    dataset_id = _create_table(api, auth_headers).json()["data"]["id"]
    ontology_id = ontology["id"]
    response = api.post(
        f"/api/v2/ontologies/{ontology_id}/mappings",
        headers=auth_headers,
        json={
            "curated_dataset_id": dataset_id,
            "entity_class": "Device",
            "field_mapping": {"编号": "code", "名称": "name", "数量": "quantity"},
            "auto_apply_on_version": True,
        },
    )
    assert response.status_code == 200, response.text
    mapping_id = response.json()["mapping_id"]

    from app.models.ontology import OntologyProject
    project = db.query(OntologyProject).filter_by(id=ontology_id).one()
    project.status = "published"
    db.commit()

    response = api.put(
        f"/api/v2/ontologies/{ontology_id}/mappings/{mapping_id}",
        headers=auth_headers,
        json={"auto_apply_on_version": False},
    )
    assert response.status_code == 409, response.text
    detail = response.json()["detail"]
    assert detail["code"] == "mapping_policy_requires_versioned_draft"
    assert detail["fields"] == ["auto_apply_on_version"]
    assert all(stage in detail["message"] for stage in ("草稿", "试跑", "发布"))

    from app.models.v2.mapping import OntologyMapping
    stored = db.query(OntologyMapping).filter_by(id=mapping_id).one()
    assert stored.field_mapping["__auto_apply_on_version__"] is True

    # Structural edits remain frozen on the same published ontology as well.
    response = api.put(
        f"/api/v2/ontologies/{ontology_id}/mappings/{mapping_id}",
        headers=auth_headers,
        json={"entity_class": "ChangedDevice"},
    )
    assert response.status_code == 409


def test_configured_upload_creates_first_version_with_field_contract(api, auth_headers):
    payload = {
        "name": "设备台账导入",
        "columns": [
            {"name": "id", "display_name": "设备编号", "type": "string", "nullable": False},
            {"name": "name", "display_name": "设备名称", "type": "string", "nullable": False},
            {"name": "quantity", "display_name": "数量", "type": "integer", "nullable": True},
        ],
        "primary_key": "id",
    }
    content = "id,name,quantity\nA1,泵机,10\nA2,阀门,5\n"
    response = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": ("原始文件.csv", io.BytesIO(content.encode()), "text/csv")},
        headers=auth_headers,
    )
    assert response.status_code == 201, response.text
    result = response.json()["data"]
    assert result["name"] == "设备台账导入"
    assert result["version_no"] == 1
    assert result["rowcount"] == 2
    assert result["source"] == "upload"

    schema_response = api.get(
        f"/api/v2/datasets/{result['id']}/schema", headers=auth_headers)
    columns = {item["name"]: item for item in schema_response.json()["columns"]}
    assert columns["id"]["display_name"] == "设备编号"
    assert columns["id"]["is_primary_key"] is True
    assert columns["name"]["nullable"] is False
    assert columns["quantity"]["type"] == "integer"


def test_configured_xlsx_upload_uses_first_sheet_and_hyperlink_display_text(
    api, auth_headers,
):
    """与前端保持首工作表语义；公式无缓存时仍能导入 HYPERLINK 显示文本。"""
    import openpyxl

    workbook = openpyxl.Workbook()
    data_sheet = workbook.active
    data_sheet.title = "待导入"
    data_sheet.append(["工单号", "说明", "数量"])
    data_sheet.append([
        '=HYPERLINK("https://12345.huawei.com/itsmnext/#/serviceDesk/viewSkillOrder/'
        'PM26072236369","PM26072236369")',
        "第一行\n第二行",
        3,
    ])
    other_sheet = workbook.create_sheet("当前打开但不导入")
    other_sheet.append(["错误表头"])
    other_sheet.append(["不应被导入"])
    workbook.active = 1
    content = io.BytesIO()
    workbook.save(content)
    workbook.close()

    payload = {
        "name": "公式工单导入",
        "columns": [
            {
                "source_key": "工单号", "name": "ticket_id",
                "display_name": "工单号", "type": "string", "nullable": False,
            },
            {
                "source_key": "说明", "name": "description",
                "display_name": "说明", "type": "string", "nullable": True,
            },
            {
                "source_key": "数量", "name": "quantity",
                "display_name": "数量", "type": "integer", "nullable": True,
            },
        ],
        "primary_key": "ticket_id",
    }
    response = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": (
            "公式工单.xlsx",
            io.BytesIO(content.getvalue()),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
        headers=auth_headers,
    )

    assert response.status_code == 201, response.text
    result = response.json()["data"]
    assert result["rowcount"] == 1
    preview = api.get(
        f"/api/v2/datasets/{result['id']}/preview", headers=auth_headers).json()
    assert preview["columns"] == ["ticket_id", "description", "quantity"]
    assert preview["rows"] == [{
        "ticket_id": "PM26072236369",
        "description": "第一行\n第二行",
        "quantity": 3,
    }]


def test_configured_xls_upload_accepts_integer_cells(
    api, auth_headers, legacy_xls_bytes,
):
    payload = {
        "name": "旧版工单导入",
        "columns": [
            {"name": "id", "type": "string", "nullable": False},
            {"name": "quantity", "type": "integer", "nullable": True},
            {"name": "note", "type": "string", "nullable": True},
        ],
        "primary_key": "id",
    }
    response = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": (
            "legacy.xls",
            io.BytesIO(legacy_xls_bytes),
            "application/vnd.ms-excel",
        )},
        headers=auth_headers,
    )

    assert response.status_code == 201, response.text
    result = response.json()["data"]
    assert result["rowcount"] == 2
    preview = api.get(
        f"/api/v2/datasets/{result['id']}/preview", headers=auth_headers).json()
    assert preview["columns"] == ["id", "quantity", "note"]
    assert preview["rows"] == [
        {"id": "A1", "quantity": 3, "note": "第一行\n第二行"},
        {"id": "A2", "quantity": 5, "note": "正常"},
    ]


def test_async_import_stages_file_and_normalizes_source_headers_to_field_keys(
    api, auth_headers, db, monkeypatch, tmp_path,
):
    """后台解析保留原始表头，正式版本只发布稳定字段标识。"""
    import openpyxl
    from sqlalchemy.orm import sessionmaker

    from app import database as database_module
    from app.config import settings
    from app.data_channel.pipeline_tasks import dispatch as dispatch_module
    from app.tasks.v2 import dataset_import as import_tasks

    monkeypatch.setattr(settings, "uploads_dir", str(tmp_path / "uploads"))
    dispatched = []
    monkeypatch.setattr(
        dispatch_module,
        "dispatch_task",
        lambda subject, payload: dispatched.append((subject, payload)),
    )
    task_session = sessionmaker(bind=db.get_bind())
    monkeypatch.setattr(database_module, "SessionLocal", task_session)

    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "默认首表"
    first.append(["设备编号", "设备名称", "数量"])
    first.append(["A1", "泵机", 10])
    second = workbook.create_sheet("不导入")
    second.append(["wrong"])
    second.append(["SHOULD-NOT-IMPORT"])
    workbook.active = 1
    content = io.BytesIO()
    workbook.save(content)
    workbook.close()

    started = api.post(
        "/api/v2/datasets/imports",
        files={"file": (
            "设备台账.xlsx",
            io.BytesIO(content.getvalue()),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
        headers=auth_headers,
    )
    assert started.status_code == 202, started.text
    job = started.json()["data"]
    assert job["status"] == "queued"
    assert dispatched == [
        ("task.dataset.import", {"job_id": job["job_id"], "kind": "inspect"}),
    ]
    import_tasks.inspect_dataset_import(job["job_id"])

    inspected = api.get(
        f"/api/v2/datasets/imports/{job['job_id']}",
        headers=auth_headers,
    )
    assert inspected.status_code == 200
    job = inspected.json()["data"]
    assert job["status"] == "ready"
    assert job["sheet_name"] == "默认首表"
    assert job["rowcount"] == 1
    assert [column["name"] for column in job["columns"]] == [
        "设备编号", "设备名称", "数量"]
    assert job["preview_rows"] == [{"设备编号": "A1", "设备名称": "泵机", "数量": 10}]

    job_dir = tmp_path / "uploads" / "dataset-imports" / job["job_id"]
    assert (job_dir / "source.xlsx").read_bytes() == content.getvalue()
    assert (job_dir / "manifest.json").is_file()
    assert (job_dir / "status.json").is_file()

    committed = api.post(
        f"/api/v2/datasets/imports/{job['job_id']}/commit",
        json={
            "name": "设备台账异步导入",
            "columns": [
                {
                    "source_key": "设备编号", "name": "device_id",
                    "display_name": "设备编号", "type": "string", "nullable": False,
                },
                {
                    "source_key": "设备名称", "name": "device_name",
                    "display_name": "设备名称", "type": "string", "nullable": False,
                },
                {
                    "source_key": "数量", "name": "quantity",
                    "display_name": "数量", "type": "integer", "nullable": True,
                },
            ],
            "primary_key": "device_id",
        },
        headers=auth_headers,
    )
    assert committed.status_code == 202, committed.text
    assert committed.json()["data"]["status"] == "import_queued"
    assert dispatched == [
        ("task.dataset.import", {"job_id": job["job_id"], "kind": "inspect"}),
        ("task.dataset.import", {"job_id": job["job_id"], "kind": "commit"}),
    ]
    import_tasks.commit_dataset_import(job["job_id"])

    completed = api.get(
        f"/api/v2/datasets/imports/{job['job_id']}",
        headers=auth_headers,
    )
    assert completed.status_code == 200
    result = completed.json()["data"]
    assert result["status"] == "completed"
    assert result["result"]["rowcount"] == 1
    assert result["result"]["source"] == "upload"

    preview = api.get(
        f"/api/v2/datasets/{result['result']['id']}/preview",
        headers=auth_headers,
    )
    assert preview.status_code == 200
    assert preview.json()["columns"] == ["device_id", "device_name", "quantity"]
    assert preview.json()["rows"] == [{
        "device_id": "A1", "device_name": "泵机", "quantity": 10,
    }]


def test_configured_upload_accepts_cr_only_csv_newlines(api, auth_headers):
    payload = {
        "name": "旧式换行 CSV",
        "columns": [
            {"name": "id", "type": "string", "nullable": False},
            {"name": "name", "type": "string", "nullable": True},
        ],
        "primary_key": "id",
    }
    response = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": (
            "old-mac.csv",
            io.BytesIO("id,name\rA1,泵机\rA2,阀门\r".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )

    assert response.status_code == 201, response.text
    assert response.json()["data"]["rowcount"] == 2


def test_configured_upload_reports_real_xlsx_parse_error(api, auth_headers):
    payload = {
        "name": "损坏表格",
        "columns": [{"name": "id", "type": "string"}],
        "primary_key": "",
    }
    response = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": (
            "broken.xlsx",
            io.BytesIO(b"PK\x03\x04not-an-xlsx\rwith-binary-data"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
        headers=auth_headers,
    )

    assert response.status_code == 400
    detail = str(response.json()["detail"])
    assert "Excel 工作簿解析失败" in detail
    assert "new-line character" not in detail


@pytest.mark.parametrize("content, expected", [
    ("id,name,quantity\nA1,,10\n", "非空列「name」"),
    ("id,name,quantity\nA1,泵机,很多\n", "类型校验未通过"),
    ("id,name,quantity\nA1,泵机,10\nA1,阀门,5\n", "主键重复"),
])
def test_configured_upload_rejects_contract_violations(api, auth_headers, content, expected):
    payload = {
        "name": "错误数据",
        "columns": [
            {"name": "id", "display_name": "编号", "type": "string", "nullable": False},
            {"name": "name", "display_name": "名称", "type": "string", "nullable": False},
            {"name": "quantity", "display_name": "数量", "type": "integer", "nullable": True},
        ],
        "primary_key": "id",
    }
    response = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": ("bad.csv", io.BytesIO(content.encode()), "text/csv")},
        headers=auth_headers,
    )
    assert response.status_code == 400
    assert expected in str(response.json()["detail"])


def test_create_table_validation(api, auth_headers):
    assert _create_table(api, auth_headers, name="  ").status_code == 400
    assert _create_table(api, auth_headers, columns=[]).status_code == 400
    assert _create_table(api, auth_headers, columns=[{"name": "  "}]).status_code == 400
    r = _create_table(api, auth_headers,
                      columns=[{"name": "编号"}, {"name": "编号"}])
    assert r.status_code == 400 and "重复" in str(r.json()["detail"])
    r = _create_table(api, auth_headers, primary_key="不存在的列")
    assert r.status_code == 400 and "不在列定义中" in str(r.json()["detail"])
    # 非法类型必须显式拒绝，不能静默改成 string 造成契约与用户选择不一致
    r = _create_table(api, auth_headers, primary_key="",
                      columns=[{"name": "a", "type": "非法类型"}, {"name": " "}])
    assert r.status_code == 400
    assert "不受支持" in str(r.json()["detail"])


# ── 在线维护 ─────────────────────────────────────────────────
def test_online_table_insert_edit_flow(api, auth_headers):
    ds_id = _create_table(api, auth_headers).json()["data"]["id"]

    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 1,
        "inserts": [
            {"values": {"编号": "A1", "名称": "泵机", "数量": "10"}},
            {"values": {"编号": "A2", "名称": "阀门", "数量": "5"}},
        ],
    }, headers=auth_headers)
    assert r.status_code == 200, r.text
    assert r.json()["version_no"] == 2 and r.json()["rowcount"] == 2

    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 2,
        "updates": [{"key": {"编号": "A1"}, "values": {"数量": "99"}}],
        "deletes": [{"key": {"编号": "A2"}}],
    }, headers=auth_headers)
    assert r.status_code == 200, r.text

    pv = api.get(f"/api/v2/datasets/{ds_id}/preview", headers=auth_headers).json()
    assert pv["total_rows"] == 1
    assert pv["rows"][0]["数量"] == "99"

    # 主键三校验照常生效：插入重复主键被拦
    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 3,
        "inserts": [{"values": {"编号": "A1", "名称": "重复", "数量": "1"}}],
    }, headers=auth_headers)
    assert r.status_code == 400 and "重复" in str(r.json()["detail"])


def test_online_table_rejects_type_violation(api, auth_headers):
    ds_id = _create_table(api, auth_headers).json()["data"]["id"]

    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 1,
        "inserts": [{"values": {"编号": "A1", "名称": "泵机", "数量": "十个"}}],
    }, headers=auth_headers)
    assert r.status_code == 400
    detail = str(r.json()["detail"])
    assert "数量" in detail and "integer" in detail
    versions = api.get(f"/api/v2/datasets/{ds_id}/versions", headers=auth_headers).json()
    assert len(versions) == 1  # 坏值没有落盘

    # "1"/"0" 在整数列合法（推断器会判 boolean，但整数录入必须放行）；空值放行
    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 1,
        "inserts": [{"values": {"编号": "A1", "名称": "", "数量": "1"}}],
    }, headers=auth_headers)
    assert r.status_code == 200, r.text


def test_online_table_declared_types_survive_edits(api, auth_headers):
    """声明类型是契约：录入整数样值后，float 列不得被重推断成 integer。"""
    r = _create_table(api, auth_headers, columns=[
        {"name": "编号", "type": "string"},
        {"name": "单价", "type": "float"},
    ])
    ds_id = r.json()["data"]["id"]

    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 1,
        "inserts": [{"values": {"编号": "A1", "单价": "5"}}],  # 看起来像 integer 的合法 float
    }, headers=auth_headers)
    assert r.status_code == 200, r.text

    sc = api.get(f"/api/v2/datasets/{ds_id}/schema", headers=auth_headers).json()
    types = {c["name"]: c["type"] for c in sc["columns"]}
    assert types == {"编号": "string", "单价": "float"}
    # 编号列声明 string：纯数字编号不得翻成 integer
    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 2,
        "inserts": [{"values": {"编号": "1001", "单价": "3.5"}}],
    }, headers=auth_headers)
    assert r.status_code == 200
    sc = api.get(f"/api/v2/datasets/{ds_id}/schema", headers=auth_headers).json()
    assert {c["name"]: c["type"] for c in sc["columns"]}["编号"] == "string"


def test_online_table_edit_rejects_unknown_column(api, auth_headers):
    ds_id = _create_table(api, auth_headers).json()["data"]["id"]
    r = api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 1,
        "inserts": [{"values": {"编号": "A1", "野列": "x"}}],
    }, headers=auth_headers)
    assert r.status_code == 400 and "不存在的列" in str(r.json()["detail"])


# ── 文件补数与列结构刷新 ──────────────────────────────────────
def test_upload_version_refreshes_declared_schema(api, auth_headers):
    ds_id = _create_table(api, auth_headers).json()["data"]["id"]

    csv = "编号,数量,备注\nB1,3,新列来自文件\n"
    r = api.post(f"/api/v2/datasets/{ds_id}/upload",
                 files={"file": ("补数.csv", io.BytesIO(csv.encode("utf-8")), "text/csv")},
                 headers=auth_headers)
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["columns_added"] == ["备注"]
    assert body["columns_removed"] == ["名称"]

    # 列结构随新文件刷新；声明过的列保留声明类型，新列按数据推断
    sc = api.get(f"/api/v2/datasets/{ds_id}/schema", headers=auth_headers).json()
    types = {c["name"]: c["type"] for c in sc["columns"]}
    assert types["数量"] == "integer" and types["编号"] == "string"
    assert "备注" in types and "名称" not in types
    pv = api.get(f"/api/v2/datasets/{ds_id}/preview", headers=auth_headers).json()
    assert pv["columns"] == ["编号", "数量", "备注"]

    # 主键契约在上传路径照常拦截：重复主键的文件不落盘
    bad = "编号,数量,备注\nB1,1,x\nB1,2,y\n"
    r = api.post(f"/api/v2/datasets/{ds_id}/upload",
                 files={"file": ("坏.csv", io.BytesIO(bad.encode("utf-8")), "text/csv")},
                 headers=auth_headers)
    assert r.status_code == 400 and "重复" in str(r.json()["detail"])


def test_upload_version_enforces_declared_column_types(api, auth_headers):
    """批量补数与在线编辑必须服从同一份类型契约，坏版本不能先落盘再告警。"""
    ds_id = _create_table(api, auth_headers).json()["data"]["id"]
    bad = "编号,名称,数量\nB1,泵机,十个\n"

    r = api.post(
        f"/api/v2/datasets/{ds_id}/upload",
        files={"file": ("类型错误.csv", io.BytesIO(bad.encode("utf-8")), "text/csv")},
        headers=auth_headers,
    )

    assert r.status_code == 400
    assert "数量" in str(r.json()["detail"])
    assert "integer" in str(r.json()["detail"])
    versions = api.get(f"/api/v2/datasets/{ds_id}/versions", headers=auth_headers).json()
    assert len(versions) == 1


# ── 本体映射准入（第一性原理：与上传数据集同权）──────────────────
def test_online_table_bindable_to_ontology_mapping(api, auth_headers, ontology):
    ds_id = _create_table(api, auth_headers).json()["data"]["id"]
    api.post(f"/api/v2/datasets/{ds_id}/rows/edit", json={
        "base_version_no": 1,
        "inserts": [{"values": {"编号": "A1", "名称": "泵机", "数量": "10"}}],
    }, headers=auth_headers)

    r = api.post(f"/api/v2/ontologies/{ontology['id']}/mappings", json={
        "curated_dataset_id": ds_id,
        "entity_class": "Item",
        "field_mapping": {"编号": "code"},
    }, headers=auth_headers)
    assert r.status_code == 200, r.text
    assert r.json()["mapping_id"]


def test_async_import_always_dispatches_nats_task(
    api, auth_headers, monkeypatch, tmp_path,
):
    """Spreadsheet imports always dispatch through the NATS work queue."""
    from app.config import settings
    from app.data_channel.pipeline_tasks import dispatch as dispatch_module

    monkeypatch.setattr(settings, "uploads_dir", str(tmp_path / "uploads"))
    dispatched = []

    monkeypatch.setattr(
        dispatch_module,
        "dispatch_task",
        lambda subject, payload: dispatched.append((subject, payload)),
    )

    started = api.post(
        "/api/v2/datasets/imports",
        files={"file": (
            "standalone.csv",
            io.BytesIO("id,name\nA1,泵机\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert started.status_code == 202, started.text
    queued = started.json()["data"]
    assert queued["status"] == "queued"
    assert queued["execution_mode"] == "nats"
    assert dispatched == [
        ("task.dataset.import", {"job_id": queued["job_id"], "kind": "inspect"}),
    ]

    inspected = api.get(
        f"/api/v2/datasets/imports/{queued['job_id']}",
        headers=auth_headers,
    )
    assert inspected.status_code == 200
    assert inspected.json()["data"]["status"] == "queued"


def test_async_import_fails_closed_when_task_channel_is_down(
    api, auth_headers, monkeypatch, tmp_path, caplog,
):
    """Task-channel failure never starts an API-process background task."""
    from app.config import settings
    from app.data_channel.pipeline_tasks import dispatch as dispatch_module

    monkeypatch.setattr(settings, "uploads_dir", str(tmp_path / "uploads"))

    def channel_down(*_args, **_kwargs):
        raise RuntimeError("nats unavailable")

    monkeypatch.setattr(dispatch_module, "dispatch_task", channel_down)
    caplog.set_level("ERROR", logger="app.data_channel.datasets.router")

    started = api.post(
        "/api/v2/datasets/imports",
        files={"file": (
            "broker-fallback.csv",
            io.BytesIO("id,name\nA1,泵机\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert started.status_code == 503, started.text
    assert started.json()["detail"] == (
        "后台任务通道不可用，数据集导入任务未投递")
    assert "任务未执行" in caplog.text
    assert "已降级" not in caplog.text
    assert "nats unavailable" not in caplog.text


def _upload_stable_manual_contract(api, auth_headers):
    payload = {
        "name": "人员台账",
        "columns": [
            {
                "source_key": "姓名",
                "name": "person_name",
                "display_name": "姓名",
                "type": "string",
                "nullable": False,
            },
            {
                "source_key": "年龄",
                "name": "age",
                "display_name": "年龄",
                "type": "integer",
                "nullable": True,
            },
        ],
        "primary_key": "person_name",
    }
    response = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": (
            "人员.csv",
            io.BytesIO("姓名,年龄\n张三,28\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert response.status_code == 201, response.text
    return response.json()["data"]["id"]


def test_declare_contract_preserves_stable_upload_types_and_mapping(
    api, auth_headers, ontology, db,
):
    payload = {
        "name": "稳定业务编码",
        "columns": [{
            "source_key": "原始编码",
            "name": "code",
            "display_name": "业务编码",
            "type": "string",
            "nullable": True,
        }],
        "primary_key": "",
    }
    uploaded = api.post(
        "/api/v2/datasets/upload",
        data={"metadata": json.dumps(payload, ensure_ascii=False)},
        files={"file": (
            "编码-v1.csv",
            io.BytesIO("原始编码\n001\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert uploaded.status_code == 201, uploaded.text
    dataset_id = uploaded.json()["data"]["id"]

    declared = api.put(
        f"/api/v2/datasets/{dataset_id}/contract",
        json={"primary_key": "code"},
        headers=auth_headers,
    )
    assert declared.status_code == 200, declared.text

    from app.models.v2.dataset import Dataset
    stored = db.query(Dataset).filter(Dataset.id == dataset_id).one()
    db.refresh(stored)
    schema = stored.schema_json
    code_column = next(
        item for item in schema["columns_typed"] if item["name"] == "code")
    assert code_column == {
        "name": "code",
        "display_name": "业务编码",
        "type": "string",
        "nullable": False,
        "source_key": "原始编码",
    }
    code_definition = next(
        item for item in schema["contract_definitions"]
        if item["field_key"] == "code")
    assert code_definition == {
        "source_key": "原始编码",
        "field_key": "code",
        "field_name": "业务编码",
        "field_type": "string",
        "is_primary_key": True,
        "nullable": False,
    }

    exposed_schema = api.get(
        f"/api/v2/datasets/{dataset_id}/schema",
        headers=auth_headers,
    )
    assert exposed_schema.status_code == 200, exposed_schema.text
    exposed_code = next(
        item for item in exposed_schema.json()["columns"]
        if item["name"] == "code")
    assert exposed_code["type"] == "string"
    assert exposed_code["display_name"] == "业务编码"
    assert exposed_code["is_primary_key"] is True
    assert exposed_code["nullable"] is False

    next_version = api.post(
        f"/api/v2/datasets/{dataset_id}/upload",
        files={"file": (
            "编码-v2.csv",
            io.BytesIO("原始编码\nABC\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert next_version.status_code == 201, next_version.text
    assert next_version.json()["version_no"] == 2

    ontology_id = ontology["id"]
    object_type = api.post(
        f"/api/v2/formal/ontologies/{ontology_id}/object-types",
        headers=auth_headers,
        json={
            "name": "StableBusinessCode",
            "displayName": "稳定业务编码",
            "primaryKey": "code",
            "properties": [{
                "id": "stable-business-code",
                "name": "code",
                "type": "string",
                "required": True,
            }],
        },
    )
    assert object_type.status_code == 201, object_type.text
    object_type_id = object_type.json()["data"]["id"]

    mapping = api.post(
        f"/api/v2/ontologies/{ontology_id}/mappings",
        headers=auth_headers,
        json={
            "curated_dataset_id": dataset_id,
            "entity_class": "StableBusinessCode",
            "target_object_type_id": object_type_id,
            "field_mapping": {"code": "code"},
        },
    )
    assert mapping.status_code == 200, mapping.text
    assert mapping.json()["mapping_id"]


def test_stable_manual_contract_normalizes_rows_and_connects_to_mapping(
    api, auth_headers, ontology, db,
):
    dataset_id = _upload_stable_manual_contract(api, auth_headers)

    preview = api.get(
        f"/api/v2/datasets/{dataset_id}/preview", headers=auth_headers).json()
    assert preview["columns"] == ["person_name", "age"]
    assert preview["rows"] == [{"person_name": "张三", "age": "28"}]

    schema = api.get(
        f"/api/v2/datasets/{dataset_id}/schema", headers=auth_headers).json()
    columns = {item["name"]: item for item in schema["columns"]}
    assert columns["person_name"]["display_name"] == "姓名"
    assert columns["person_name"]["is_primary_key"] is True
    assert columns["age"]["display_name"] == "年龄"
    assert columns["age"]["type"] == "integer"

    from app.models.v2.dataset import Dataset
    stored = db.query(Dataset).filter(Dataset.id == dataset_id).one()
    assert stored.schema_json["manual_field_contract_version"] == 2
    assert stored.schema_json["contract_definitions"] == [
        {
            "source_key": "姓名",
            "field_key": "person_name",
            "field_name": "姓名",
            "field_type": "string",
            "is_primary_key": True,
            "nullable": False,
        },
        {
            "source_key": "年龄",
            "field_key": "age",
            "field_name": "年龄",
            "field_type": "integer",
            "is_primary_key": False,
            "nullable": True,
        },
    ]

    mapping = api.post(
        f"/api/v2/ontologies/{ontology['id']}/mappings",
        headers=auth_headers,
        json={
            "curated_dataset_id": dataset_id,
            "entity_class": "Person",
            "field_mapping": {
                "person_name": "name",
                "age": "age",
            },
        },
    )
    assert mapping.status_code == 200, mapping.text
    assert mapping.json()["mapping_id"]


def test_stable_manual_contract_locks_follow_up_version_headers(
    api, auth_headers,
):
    dataset_id = _upload_stable_manual_contract(api, auth_headers)

    source_header = api.post(
        f"/api/v2/datasets/{dataset_id}/upload",
        files={"file": (
            "人员-v2.csv",
            io.BytesIO("姓名,年龄\n李四,31\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert source_header.status_code == 201, source_header.text
    preview = api.get(
        f"/api/v2/datasets/{dataset_id}/preview", headers=auth_headers).json()
    assert preview["rows"] == [{"person_name": "李四", "age": "31"}]

    field_header = api.post(
        f"/api/v2/datasets/{dataset_id}/upload",
        files={"file": (
            "人员-v3.csv",
            io.BytesIO("person_name,age\n王五,42\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert field_header.status_code == 201, field_header.text
    preview = api.get(
        f"/api/v2/datasets/{dataset_id}/preview", headers=auth_headers).json()
    assert preview["rows"] == [{"person_name": "王五", "age": "42"}]

    before_versions = api.get(
        f"/api/v2/datasets/{dataset_id}/versions", headers=auth_headers).json()
    mixed = api.post(
        f"/api/v2/datasets/{dataset_id}/upload",
        files={"file": (
            "人员-混用.csv",
            io.BytesIO("姓名,age\n赵六,35\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert mixed.status_code == 400
    assert "不允许混用、新增、缺失或调整字段顺序" in str(
        mixed.json()["detail"])
    after_versions = api.get(
        f"/api/v2/datasets/{dataset_id}/versions", headers=auth_headers).json()
    assert len(after_versions) == len(before_versions)

    empty = api.post(
        f"/api/v2/datasets/{dataset_id}/upload",
        files={"file": (
            "人员-空表.csv",
            io.BytesIO("姓名,年龄\n".encode("utf-8")),
            "text/csv",
        )},
        headers=auth_headers,
    )
    assert empty.status_code == 201, empty.text
    preview = api.get(
        f"/api/v2/datasets/{dataset_id}/preview", headers=auth_headers).json()
    assert preview["columns"] == ["person_name", "age"]
    assert preview["rows"] == []


@pytest.mark.parametrize("field_key", ["姓名", "PersonName", "_name", "person-name"])
def test_stable_manual_contract_rejects_invalid_field_identifier(
    api, auth_headers, field_key,
):
    response = api.post(
        "/api/v2/datasets/create-table",
        headers=auth_headers,
        json={
            "name": "不合法字段",
            "columns": [{
                "source_key": field_key,
                "name": field_key,
                "display_name": "姓名",
                "type": "string",
            }],
        },
    )
    assert response.status_code == 400
    assert "必须以小写字母开头" in str(response.json()["detail"])
