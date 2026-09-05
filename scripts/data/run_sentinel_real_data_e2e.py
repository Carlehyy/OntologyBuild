#!/usr/bin/env python3
"""Run the Sentinel production path against a live API and real supplier data.

This is intentionally not a pytest/TestClient test.  It talks to a running
OpenOntology backend, uploads the repository's real supplier workbook, goes
through draft -> isolated trial -> promotion, publishes a second immutable
dataset version, and verifies the resulting CDC/Sentinel/action cascade.
"""
from __future__ import annotations

import argparse
import io
import json
import sys
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DATA = ROOT / "test_data" / "供应链" / "supplier_database.xlsx"
TARGET_SUPPLIER_ID = "SUP005"
TARGET_NEW_ONTIME_RATE = 86.4


class CheckFailed(RuntimeError):
    pass


def require(condition: bool, message: str) -> None:
    if not condition:
        raise CheckFailed(message)


def unwrap(payload: Any) -> Any:
    if isinstance(payload, dict) and "data" in payload:
        return payload["data"]
    return payload


class Api:
    def __init__(self, base_url: str) -> None:
        self.client = httpx.Client(base_url=base_url.rstrip("/"), timeout=90.0)
        self.headers: dict[str, str] = {}

    def close(self) -> None:
        self.client.close()

    def request(self, method: str, path: str, **kwargs: Any) -> Any:
        headers = dict(self.headers)
        headers.update(kwargs.pop("headers", {}))
        response = self.client.request(method, path, headers=headers, **kwargs)
        if not response.is_success:
            raise CheckFailed(
                f"{method} {path} failed ({response.status_code}): "
                f"{response.text[:3000]}"
            )
        if response.status_code == 204:
            return None
        return unwrap(response.json())

    def login(self, username: str, password: str) -> None:
        data = self.request(
            "POST", "/api/v1/auth/login",
            json={"username": username, "password": password},
        )
        token = data.get("access_token")
        require(bool(token), "login response did not contain access_token")
        self.headers = {"Authorization": f"Bearer {token}"}


def workbook_with_new_rate(source: Path) -> tuple[bytes, str]:
    workbook = load_workbook(source)
    sheet = workbook.active
    headers = {
        str(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value is not None
    }
    require("供应商ID" in headers, "supplier workbook is missing 供应商ID")
    require("供应商名称" in headers, "supplier workbook is missing 供应商名称")
    require("准时率%" in headers, "supplier workbook is missing 准时率%")

    supplier_name = ""
    changed = False
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row, headers["供应商ID"]).value) != TARGET_SUPPLIER_ID:
            continue
        supplier_name = str(sheet.cell(row, headers["供应商名称"]).value)
        old_rate = float(sheet.cell(row, headers["准时率%"]).value)
        require(
            old_rate >= 90,
            f"{TARGET_SUPPLIER_ID} must start outside the Sentinel condition",
        )
        sheet.cell(row, headers["准时率%"]).value = TARGET_NEW_ONTIME_RATE
        changed = True
        break
    require(changed, f"supplier {TARGET_SUPPLIER_ID} was not found")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), supplier_name


def create_table(api: Api, name: str) -> str:
    columns = [
        ("供应商ID", "string"),
        ("供应商名称", "string"),
        ("等级", "string"),
        ("主供物料", "string"),
        ("年采购额(万)", "float"),
        ("准时率%", "float"),
        ("合格率%", "float"),
        ("联系人", "string"),
        ("状态", "string"),
    ]
    dataset = api.request(
        "POST", "/api/v2/datasets/create-table",
        json={
            "name": name,
            "columns": [
                {
                    "name": column,
                    "display_name": column,
                    "type": column_type,
                    # The real workbook contains legitimate blanks outside its
                    # business key (for example SUP005 合格率%). Preserve that
                    # source contract instead of manufacturing test-only data.
                    "nullable": column == "合格率%",
                }
                for column, column_type in columns
            ],
            "primary_key": "供应商ID",
        },
    )
    return str(dataset["id"])


def upload_workbook(api: Api, dataset_id: str, filename: str, content: bytes) -> Any:
    return api.request(
        "POST", f"/api/v2/datasets/{dataset_id}/upload",
        files={
            "file": (
                filename,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def by_property(rows: list[dict], property_name: str) -> dict[str, dict]:
    result: dict[str, dict] = {}
    for row in rows:
        value = (row.get("properties") or {}).get(property_name)
        if value is not None:
            result[str(value)] = row
    return result


def assert_dispatch_success(result: dict) -> None:
    require(
        result.get("projection_warnings") in (None, []),
        f"mapping query projection degraded: "
        f"{result.get('projection_warnings')}",
    )
    require(
        result.get("neo4j_projection_rebuilt") is True,
        f"Neo4j projection was not rebuilt: {result}",
    )
    dispatch = result.get("sentinel_dispatch")
    require(isinstance(dispatch, dict), "build-all did not report sentinel_dispatch")
    errors = dispatch.get("errors")
    require(errors in (None, 0, []), f"sentinel dispatch errors: {errors}")
    for run in dispatch.get("runs") or []:
        nested = run.get("result") or {}
        nested_errors = nested.get("errors")
        require(
            nested_errors in (None, 0, []),
            f"nested Sentinel evaluation failed: {nested}",
        )


def wait_for_cdc(
    api: Api,
    ontology_id: str,
    *,
    timeout_seconds: float = 90.0,
) -> dict:
    """Wait until the durable CDC consumer has finished this ontology's work.

    A retry is an observable, recoverable delivery state rather than data loss.
    Projection and control events can briefly race at release/enable boundaries;
    the production contract is that the durable consumer eventually becomes
    healthy and quiescent.  Dead letters, a missing worker, or timeout remain
    hard failures.
    """
    deadline = time.monotonic() + timeout_seconds
    last_status: dict = {}
    while time.monotonic() < deadline:
        status = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/cdc-status",
        )
        require(
            isinstance(status, dict),
            f"CDC status response is not an object: {status!r}",
        )
        last_status = status
        durable = status.get("durable") or {}
        require(
            status.get("worker_alive") is True,
            f"durable Sentinel CDC worker is unavailable: {status}",
        )
        require(
            int(durable.get("dead") or 0) == 0,
            f"durable Sentinel CDC contains dead letters: {status}",
        )
        if (
            status.get("healthy") is True
            and status.get("quiescent") is True
        ):
            return status
        time.sleep(0.2)
    raise CheckFailed(
        f"durable Sentinel CDC did not recover and become quiescent in "
        f"{timeout_seconds:.0f}s: {last_status}"
    )


def run(args: argparse.Namespace) -> dict[str, Any]:
    data_file = Path(args.data_file).expanduser().resolve()
    require(data_file.is_file(), f"real data file does not exist: {data_file}")
    original_bytes = data_file.read_bytes()
    changed_bytes, target_supplier_name = workbook_with_new_rate(data_file)

    api = Api(args.base_url)
    try:
        api.login(args.username, args.password)
        suffix = f"{int(time.time() * 1000):x}"
        supplier_type_id = f"ot-real-supplier-{suffix}"
        alert_type_id = f"ot-real-alert-{suffix}"
        risk_function_id = f"fn-supplier-risk-{suffix}"
        create_alert_action_id = f"act-create-alert-{suffix}"
        notify_action_id = f"act-notify-alert-{suffix}"
        dynamic_notify_action_id = f"act-dynamic-notify-{suffix}"
        risk_sentinel_id = f"sentinel-risk-{suffix}"
        trial_probe_sentinel_id = f"sentinel-trial-probe-{suffix}"
        alert_sentinel_id = f"sentinel-alert-{suffix}"

        ontology = api.request(
            "POST", "/api/v1/ontologies",
            json={
                "name": f"哨兵真实供应链闭环-{suffix}",
                "domain": "供应链",
                "description": "Sentinel production-path real-data verification",
            },
        )
        ontology_id = str(ontology["id"])
        tree = api.request(
            "GET", f"/api/v2/ontologies/{ontology_id}/version-tree")
        root = next(
            item for item in tree["versions"]
            if item["version_number"] == "v0"
        )
        draft = api.request(
            "POST",
            f"/api/v2/ontologies/{ontology_id}/versions/{root['id']}/drafts",
            json={
                "versionLabel": "真实供应商哨兵验证",
                "description": "真实数据隔离试跑、发布和增量级联",
            },
        )
        require(
            draft.get("node_kind") == "draft"
            and draft.get("lifecycle_status") == "editing",
            f"new version did not enter draft/editing state: {draft}",
        )

        supplier_properties = [
            ("supplier-id", "供应商ID", "string"),
            ("supplier-name", "供应商名称", "string"),
            ("supplier-level", "等级", "string"),
            ("supplier-material", "主供物料", "string"),
            ("supplier-spend", "年采购额(万)", "number"),
            ("supplier-ontime", "准时率%", "number"),
            ("supplier-quality", "合格率%", "number"),
            ("supplier-contact", "联系人", "string"),
            ("supplier-status", "状态", "string"),
        ]
        workspace = {
            "baseRevision": f"{draft['revision']}:{draft['snapshot_hash']}",
            "version": draft["version_number"],
            "objectTypes": [
                {
                    "id": supplier_type_id,
                    "name": "RealSupplier",
                    "displayName": "真实供应商",
                    "primaryKey": "supplier-id",
                    "properties": [
                        {
                            "id": property_id,
                            "name": name,
                            "displayName": name,
                            "type": property_type,
                            "required": name != "合格率%",
                            "source": "stored",
                        }
                        for property_id, name, property_type in supplier_properties
                    ] + [{
                        "id": "supplier-risk-score",
                        "name": "risk_score",
                        "displayName": "准时风险分",
                        "type": "number",
                        "required": False,
                        "source": "computed",
                        "computed": True,
                        "functionId": risk_function_id,
                    }],
                },
                {
                    "id": alert_type_id,
                    "name": "SupplierAlert",
                    "displayName": "供应商告警",
                    "primaryKey": "alert-id",
                    "properties": [
                        {
                            "id": "alert-id", "name": "alert_id",
                            "displayName": "告警编号", "type": "string",
                            "required": True, "source": "stored",
                        },
                        {
                            "id": "alert-supplier-name", "name": "supplier_name",
                            "displayName": "供应商名称", "type": "string",
                            "required": True, "source": "stored",
                        },
                        {
                            "id": "alert-status", "name": "status",
                            "displayName": "告警状态", "type": "string",
                            "required": True, "source": "stored",
                        },
                    ],
                },
            ],
            "linkTypes": [],
            "actions": [
                {
                    "id": create_alert_action_id,
                    "name": "create_supplier_alert",
                    "displayName": "创建供应商告警",
                    "objectTypeId": supplier_type_id,
                    # Exercise the real HITL checkpoint/resume path as part of
                    # the production Sentinel chain, not only in unit tests.
                    "requiresApproval": True,
                    "parameters": [
                        {
                            "id": "alert-supplier-name-param",
                            "name": "supplier_name",
                            "displayName": "供应商名称",
                            "type": "string",
                            "required": True,
                        }
                    ],
                    "rules": [
                        {
                            "id": "create-alert-object",
                            "type": "create_object",
                            "name": "真实创建告警对象",
                            "enabled": True,
                            "order": 0,
                            "config": {
                                "targetObjectTypeId": alert_type_id,
                                "propertyMappings": [
                                    {
                                        "targetProperty": "supplier_name",
                                        "sourceType": "parameter",
                                        "sourceValue": "supplier_name",
                                    },
                                    {
                                        "targetProperty": "status",
                                        "sourceType": "constant",
                                        "sourceValue": "\"open\"",
                                    },
                                ],
                            },
                        }
                    ],
                },
                {
                    "id": notify_action_id,
                    "name": "notify_supplier_alert",
                    "displayName": "发送供应商告警通知",
                    "objectTypeId": alert_type_id,
                    "requiresApproval": False,
                    "parameters": [
                        {
                            "id": "notify-supplier-name-param",
                            "name": "supplier_name",
                            "displayName": "供应商名称",
                            "type": "string",
                            "required": True,
                        }
                    ],
                    "rules": [
                        {
                            "id": "send-internal-alert",
                            "type": "notification",
                            "name": "发送真实站内通知",
                            "enabled": True,
                            "order": 0,
                            "config": {
                                "channel": "internal",
                                "recipientSource": "constant",
                                "recipient": "admin",
                                "subject": "供应商准时率告警",
                                "messageTemplate": (
                                    "供应商 {{params.supplier_name}} "
                                    "准时率跌破 90%"
                                ),
                            },
                        }
                    ],
                },
                {
                    "id": dynamic_notify_action_id,
                    "name": "notify_dynamic_supplier_risk",
                    "displayName": "动态哨兵供应商风险通知",
                    "objectTypeId": supplier_type_id,
                    "requiresApproval": False,
                    "parameters": [
                        {
                            "id": "dynamic-supplier-name-param",
                            "name": "supplier_name",
                            "displayName": "供应商名称",
                            "type": "string",
                            "required": True,
                        }
                    ],
                    "rules": [
                        {
                            "id": "send-dynamic-internal-alert",
                            "type": "notification",
                            "name": "发送动态哨兵真实通知",
                            "enabled": True,
                            "order": 0,
                            "config": {
                                "channel": "internal",
                                "recipientSource": "constant",
                                "recipient": "admin",
                                "subject": "动态哨兵供应商风险",
                                "messageTemplate": (
                                    "动态哨兵发现供应商 "
                                    "{{params.supplier_name}} 准时风险升高"
                                ),
                            },
                        }
                    ],
                },
            ],
            "functions": [
                {
                    "id": risk_function_id,
                    "name": "supplier_risk_score",
                    "displayName": "计算供应商准时风险分",
                    "functionType": "object",
                    "language": "expression",
                    "targetObjectTypeId": supplier_type_id,
                    "parameters": [],
                    "returnType": "number",
                    "body": "100 - object['准时率%']",
                    "enabled": True,
                }
            ],
            "instances": [],
            "linkInstances": [],
        }
        saved = api.request(
            "PUT",
            f"/api/v2/ontologies/{ontology_id}/versions/{draft['id']}/workspace",
            json=workspace,
        )

        supplier_dataset_id = create_table(
            api, f"真实供应商主数据-{suffix}")
        alert_seed_dataset_id = create_table(
            api, f"真实供应商告警基线-{suffix}")
        upload_workbook(
            api, supplier_dataset_id, data_file.name, original_bytes)
        upload_workbook(
            api, alert_seed_dataset_id, data_file.name, original_bytes)

        supplier_field_mapping = {
            name: name for _, name, _ in supplier_properties
        }
        supplier_field_mapping["__primary_key__"] = "供应商ID"
        # This is an explicit production subscription: publishing a version
        # means later immutable dataset versions are automatically reconciled.
        # The production gate must reject a manual mapping without this flag.
        supplier_field_mapping["__auto_apply_on_version__"] = True
        mapping_saved = api.request(
            "PUT",
            (
                f"/api/v2/ontologies/{ontology_id}/versions/"
                f"{draft['id']}/workspace/mappings"
            ),
            json={
                "baseRevision": saved["revision"],
                "mappings": [
                    {
                        "id": f"map-real-supplier-{suffix}",
                        "curatedDatasetId": supplier_dataset_id,
                        "entityClass": "RealSupplier",
                        "targetObjectTypeId": supplier_type_id,
                        "fieldMapping": supplier_field_mapping,
                        "status": "draft",
                        "confidence": 1,
                    },
                    {
                        "id": f"map-real-alert-seed-{suffix}",
                        "curatedDatasetId": alert_seed_dataset_id,
                        "entityClass": "SupplierAlert",
                        "targetObjectTypeId": alert_type_id,
                        "fieldMapping": {
                            "供应商ID": "alert_id",
                            "供应商名称": "supplier_name",
                            "状态": "status",
                            "__primary_key__": "供应商ID",
                            "__auto_apply_on_version__": True,
                        },
                        "status": "draft",
                        "confidence": 1,
                    },
                ],
                "linkMappings": [],
                "sentinels": [
                    {
                        "id": risk_sentinel_id,
                        "name": "real_supplier_ontime_risk",
                        "displayName": "真实供应商准时率风险",
                        "bindings": [
                            {
                                "alias": "supplier",
                                "objectTypeId": supplier_type_id,
                                "filter": None,
                            }
                        ],
                        "links": [],
                        "condition": (
                            "supplier['供应商ID'] == 'SUP005' and "
                            "supplier.risk_score > 10"
                        ),
                        "primaryAlias": "supplier",
                        "actionIds": [create_alert_action_id],
                        "actionParameters": {
                            create_alert_action_id: {
                                "supplier_name": {
                                    "sourceType": "property",
                                    "alias": "supplier",
                                    "property": "供应商名称",
                                }
                            }
                        },
                        "onChange": True,
                        "onSchedule": False,
                        "scanIntervalSeconds": 300,
                        "triggerMode": "on_enter",
                        "muted": False,
                        "enabled": True,
                    },
                    {
                        # Disabled in the release, but trial must still validate
                        # and dry-plan it. This catches latent parameter/action
                        # failures before a later operational enable.
                        "id": trial_probe_sentinel_id,
                        "name": "real_disabled_trial_probe",
                        "displayName": "禁用哨兵试跑探针",
                        "bindings": [
                            {
                                "alias": "supplier",
                                "objectTypeId": supplier_type_id,
                                "filter": None,
                            }
                        ],
                        "links": [],
                        "condition": (
                            "supplier['供应商ID'] == 'SUP007' and "
                            "supplier.risk_score > 10"
                        ),
                        "primaryAlias": "supplier",
                        "actionIds": [create_alert_action_id],
                        "actionParameters": {
                            create_alert_action_id: {
                                "supplier_name": {
                                    "sourceType": "property",
                                    "alias": "supplier",
                                    "property": "供应商名称",
                                }
                            }
                        },
                        "onChange": True,
                        "onSchedule": False,
                        "scanIntervalSeconds": 300,
                        "triggerMode": "on_enter",
                        "muted": False,
                        "enabled": False,
                    },
                    {
                        "id": alert_sentinel_id,
                        "name": "real_supplier_alert_created",
                        "displayName": "真实供应商告警已创建",
                        "bindings": [
                            {
                                "alias": "alert",
                                "objectTypeId": alert_type_id,
                                "filter": None,
                            }
                        ],
                        "links": [],
                        "condition": "alert.status == 'open'",
                        "primaryAlias": "alert",
                        "actionIds": [notify_action_id],
                        "actionParameters": {
                            notify_action_id: {
                                "supplier_name": {
                                    "sourceType": "property",
                                    "alias": "alert",
                                    "property": "supplier_name",
                                }
                            }
                        },
                        "onChange": True,
                        "onSchedule": False,
                        "scanIntervalSeconds": 300,
                        "triggerMode": "on_enter",
                        "muted": False,
                        "enabled": True,
                    },
                ],
            },
        )
        require(bool(mapping_saved.get("snapshotHash")), "mapping save failed")

        trial = api.request(
            "POST",
            (
                f"/api/v2/ontologies/{ontology_id}/versions/"
                f"{draft['id']}/trial-runs"
            ),
            json={},
        )
        require(
            trial["status"] == "passed",
            f"real-data isolated trial did not pass: {trial}",
        )
        trial_version = api.request(
            "GET",
            f"/api/v2/ontologies/{ontology_id}/versions/{draft['id']}",
        )
        require(
            trial_version.get("node_kind") == "draft"
            and trial_version.get("lifecycle_status") == "trial_ready",
            f"passed trial did not freeze the draft in trial_ready: "
            f"{trial_version}",
        )
        trial_result = trial.get("result") or {}
        require(
            trial_result.get("actionsExecuted") == 0,
            "isolated trial executed a real side effect",
        )
        require(
            trial_result.get("sideEffects") == "blocked",
            "isolated trial did not report side effects as blocked",
        )
        trial_suppliers = by_property(
            (trial_result.get("samples") or {}).get("objects") or [],
            "供应商ID",
        )
        trial_target = trial_suppliers.get(TARGET_SUPPLIER_ID)
        require(
            trial_target is not None,
            "isolated trial did not materialize the target real supplier",
        )
        require(
            abs(float((trial_target.get("computed") or {})["risk_score"]) - 6.8)
            < 0.000001,
            f"isolated trial computed value drifted: {trial_target}",
        )
        trial_sentinels = {
            str(item.get("id")): item
            for item in trial_result.get("sentinels") or []
        }
        probe = trial_sentinels.get(trial_probe_sentinel_id) or {}
        require(
            probe.get("matched") == 1
            and probe.get("plannedActions") == 1,
            "disabled built-in Sentinel was not fully action-previewed in trial: "
            f"{probe}",
        )
        require(
            probe.get("sideEffects") == "none",
            f"trial action preview reported a side effect: {probe}",
        )

        impact = api.request(
            "GET",
            f"/api/v2/ontologies/{ontology_id}/versions/{draft['id']}/impact",
        )
        readiness = impact.get("releaseReadiness") or {}
        require(
            readiness.get("ready") is True,
            f"release readiness failed: {readiness}",
        )
        release = api.request(
            "POST",
            (
                f"/api/v2/ontologies/{ontology_id}/versions/"
                f"{draft['id']}/promote"
            ),
            json={
                "trialRunId": trial["id"],
                "impactHash": impact["impactHash"],
                "versionLabel": "真实供应商哨兵发布",
            },
        )
        release_id = str(release["id"])
        require(
            release.get("node_kind") == "release"
            and release.get("lifecycle_status") == "released",
            f"promotion did not create an immutable release: {release}",
        )
        query_projection = release.get("query_projection") or {}
        require(
            query_projection.get("ready") is True
            and query_projection.get("neo4j") == "ok",
            f"promotion query projections were not ready: {query_projection}",
        )
        superseded_draft = api.request(
            "GET",
            f"/api/v2/ontologies/{ontology_id}/versions/{draft['id']}",
        )
        require(
            superseded_draft.get("lifecycle_status") == "superseded",
            f"promoted draft remained mutable/active: {superseded_draft}",
        )
        # Promotion materializes the released projection and can itself enqueue
        # durable CDC work.  Drain it before publishing the changed workbook so
        # an old root event cannot observe the newer state and blur causality.
        wait_for_cdc(api, ontology_id)

        supplier_instances_path = (
            f"/api/v2/formal/ontologies/{ontology_id}/instances"
            f"?object_type_id={supplier_type_id}"
            f"&expected_release_id={release_id}"
        )
        alert_instances_path = (
            f"/api/v2/formal/ontologies/{ontology_id}/instances"
            f"?object_type_id={alert_type_id}"
            f"&expected_release_id={release_id}"
        )
        initial_suppliers = api.request("GET", supplier_instances_path)
        initial_alerts = api.request("GET", alert_instances_path)
        initial_supplier_ids = {
            key: row["id"]
            for key, row in by_property(
                initial_suppliers, "供应商ID").items()
        }
        initial_seed_alert_ids = {
            key: row["id"]
            for key, row in by_property(
                initial_alerts, "alert_id").items()
            if row.get("source") != "action"
        }
        require(
            len(initial_supplier_ids) == 10,
            f"expected 10 promoted suppliers, got {len(initial_supplier_ids)}",
        )
        require(
            len(initial_seed_alert_ids) == 10,
            f"expected 10 promoted alert seeds, got {len(initial_seed_alert_ids)}",
        )
        require(
            not [
                row for row in initial_alerts
                if row.get("source") == "action"
            ],
            "trial or promotion unexpectedly executed the create-object action",
        )

        # The second built-in was intentionally released disabled while its
        # trial preview proved that SUP007 already matches. Enabling the exact
        # immutable-release member must therefore create one durable
        # initialization event and execute the pre-existing match; waiting for
        # some unrelated future data change would make "enabled" misleading.
        released_sentinels = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/",
        )
        released_by_id = {
            str(item.get("id")): item for item in released_sentinels
        }
        probe_overlay = released_by_id.get(trial_probe_sentinel_id) or {}
        require(
            probe_overlay.get("enabled") is False
            and probe_overlay.get("releaseId") == release_id
            and isinstance(probe_overlay.get("enableGeneration"), int),
            f"disabled built-in operational lineage is incomplete: "
            f"{probe_overlay}",
        )
        enabled_probe = api.request(
            "PATCH",
            (
                f"/api/v1/ontologies/{ontology_id}/sentinels/"
                f"{trial_probe_sentinel_id}/operational-state"
            ),
            json={
                "enabled": True,
                "expectedReleaseId": release_id,
                "expectedGeneration": probe_overlay["enableGeneration"],
            },
        )
        require(
            enabled_probe.get("enabled") is True
            and enabled_probe.get("enableGeneration")
            == probe_overlay["enableGeneration"] + 1,
            f"built-in Sentinel operational activation failed: "
            f"{enabled_probe}",
        )
        wait_for_cdc(api, ontology_id)
        initial_supplier_map = by_property(
            initial_suppliers, "供应商ID")
        probe_supplier_name = str(
            (initial_supplier_map.get("SUP007") or {}).get(
                "properties", {}
            ).get("供应商名称") or ""
        )
        require(
            bool(probe_supplier_name),
            "real supplier SUP007 is missing from the released projection",
        )
        probe_pending = api.request(
            "GET",
            (
                f"/api/v2/formal/ontologies/{ontology_id}/pending-actions"
                f"?release_id={release_id}"
            ),
        )
        probe_proposals = [
            item for item in probe_pending
            if item.get("actionId") == create_alert_action_id
            and (item.get("parameters") or {}).get("supplier_name")
            == probe_supplier_name
        ]
        require(
            len(probe_proposals) == 1,
            "enabling the built-in Sentinel did not produce exactly one "
            f"durable action proposal for its existing match: {probe_proposals}",
        )
        probe_approval = api.request(
            "POST",
            (
                f"/api/v2/formal/ontologies/{ontology_id}/action-logs/"
                f"{probe_proposals[0]['id']}/decide"
            ),
            json={
                "decision": "approved",
                "reason": "Codex 内建哨兵启用真实数据 E2E 审批",
                "releaseId": release_id,
            },
        )
        require(
            (probe_approval.get("executionLog") or {}).get("status")
            == "success"
            and (probe_approval.get("sentinelResume") or {}).get("status")
            == "fired",
            f"built-in Sentinel activation did not execute after approval: "
            f"{probe_approval}",
        )
        wait_for_cdc(api, ontology_id)

        api.request(
            "PUT",
            f"/api/v2/formal/ontologies/{ontology_id}/agent/profile",
            json={"allowedActionIds": [dynamic_notify_action_id]},
        )
        dynamic = api.request(
            "POST",
            (
                f"/api/v2/formal/ontologies/{ontology_id}"
                "/agent/dynamic-sentinels"
            ),
            json={
                "releaseId": release_id,
                "definition": {
                    "name": f"assistant_supplier_risk_{suffix}",
                    "displayName": "助手动态供应商风险",
                    "description": "真实数据动态哨兵执行验证",
                    "bindings": [{
                        "alias": "supplier",
                        "objectTypeId": supplier_type_id,
                        "filter": None,
                    }],
                    "links": [],
                    "condition": (
                        "supplier['供应商ID'] == 'SUP005' and "
                        "supplier.risk_score > 5"
                    ),
                    "conditionRows": [],
                    "conditionLogic": "and",
                    "primaryAlias": "supplier",
                    "actionIds": [dynamic_notify_action_id],
                    "actionParameters": {
                        dynamic_notify_action_id: {
                            "supplier_name": {
                                "sourceType": "property",
                                "alias": "supplier",
                                "property": "供应商名称",
                            }
                        }
                    },
                    "onChange": True,
                    "onSchedule": False,
                    "scanIntervalSeconds": 300,
                    "triggerMode": "on_enter",
                    "muted": False,
                },
            },
        )
        dynamic_sentinel_id = str(dynamic["id"])
        require(
            dynamic.get("origin") == "assistant_dynamic"
            and dynamic.get("enabled") is False,
            f"assistant Sentinel governance defaults are unsafe: {dynamic}",
        )
        dynamic_trial = api.request(
            "POST",
            (
                f"/api/v2/formal/ontologies/{ontology_id}"
                f"/agent/dynamic-sentinels/{dynamic_sentinel_id}/trial"
            ),
            json={"releaseId": release_id},
        )
        dynamic_report = dynamic_trial.get("lastTrialReport") or {}
        require(
            dynamic_report.get("passed") is True
            and dynamic_report.get("matchCount") == 1
            and dynamic_report.get("plannedActionCount") == 1
            and dynamic_report.get("sideEffects") == "none",
            f"assistant Sentinel isolated trial was not deterministic: "
            f"{dynamic_report}",
        )
        dynamic_plans = dynamic_report.get("plannedActions") or []
        require(
            len(dynamic_plans) == 1
            and dynamic_plans[0].get("status") == "success"
            and (dynamic_plans[0].get("parameters") or {}).get(
                "supplier_name"
            ) == target_supplier_name,
            f"assistant Sentinel trial did not resolve its real action "
            f"parameter: {dynamic_plans}",
        )
        enabled_dynamic = api.request(
            "POST",
            (
                f"/api/v2/formal/ontologies/{ontology_id}"
                f"/agent/dynamic-sentinels/{dynamic_sentinel_id}/enabled"
            ),
            json={
                "releaseId": release_id,
                "expectedRevision": dynamic_trial["definitionRevision"],
                "enabled": True,
            },
        )
        require(
            enabled_dynamic.get("enabled") is True,
            "assistant Sentinel did not become executable after current trial",
        )
        wait_for_cdc(api, ontology_id)
        activation_firings = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/firings?limit=100",
        )
        activation_notifications = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/notifications?limit=100",
        )
        require(
            len([
                item for item in activation_firings
                if item.get("sentinelId") == dynamic_sentinel_id
                and item.get("status") == "fired"
            ]) == 1,
            "assistant Sentinel enable did not durably initialize the existing "
            "trial-proven match",
        )
        require(
            len([
                item for item in activation_notifications
                if item.get("actionId") == dynamic_notify_action_id
                and target_supplier_name in str(item.get("body") or "")
            ]) == 1,
            "assistant Sentinel activation did not execute its real action "
            "exactly once",
        )

        upload_workbook(
            api,
            supplier_dataset_id,
            "supplier_database_rate_drop.xlsx",
            changed_bytes,
        )
        first_build = api.request(
            "POST",
            f"/api/v2/ontologies/{ontology_id}/mappings/build-all",
            json={},
        )
        assert_dispatch_success(first_build)
        wait_for_cdc(api, ontology_id)

        pending_actions = api.request(
            "GET",
            (
                f"/api/v2/formal/ontologies/{ontology_id}/pending-actions"
                f"?release_id={release_id}"
            ),
        )
        pending_create_actions = [
            item for item in pending_actions
            if item.get("actionId") == create_alert_action_id
        ]
        require(
            len(pending_create_actions) == 1,
            "real-data Sentinel did not create exactly one durable HITL "
            f"proposal: {pending_create_actions}",
        )
        pending_create = pending_create_actions[0]
        require(
            pending_create.get("status") == "pending"
            and pending_create.get("ontologyReleaseId") == release_id
            and pending_create.get("sentinelMatchStateId")
            and (pending_create.get("parameters") or {}).get("supplier_name")
            == target_supplier_name,
            f"HITL proposal lost release, match, or parameter lineage: "
            f"{pending_create}",
        )
        approval = api.request(
            "POST",
            (
                f"/api/v2/formal/ontologies/{ontology_id}/action-logs/"
                f"{pending_create['id']}/decide"
            ),
            json={
                "decision": "approved",
                "reason": "Codex 真实供应链 E2E 审批",
                "releaseId": release_id,
            },
        )
        require(
            (approval.get("pendingLog") or {}).get("status") == "approved"
            and (approval.get("executionLog") or {}).get("status") == "success"
            and approval.get("decisionFactId")
            and (approval.get("sentinelResume") or {}).get("status") == "fired",
            f"HITL approval did not durably execute and resume Sentinel: "
            f"{approval}",
        )
        wait_for_cdc(api, ontology_id)

        updated_suppliers = api.request("GET", supplier_instances_path)
        updated_alerts = api.request("GET", alert_instances_path)
        updated_supplier_map = by_property(
            updated_suppliers, "供应商ID")
        require(
            {
                key: row["id"] for key, row in updated_supplier_map.items()
            } == initial_supplier_ids,
            "Supplier object IDs changed between promotion and live reprojection",
        )
        updated_seed_alert_ids = {
            key: row["id"]
            for key, row in by_property(updated_alerts, "alert_id").items()
            if row.get("source") != "action"
        }
        require(
            updated_seed_alert_ids == initial_seed_alert_ids,
            "Alert seed object IDs changed between promotion and live reprojection",
        )
        target = updated_supplier_map.get(TARGET_SUPPLIER_ID)
        require(target is not None, f"{TARGET_SUPPLIER_ID} disappeared after build")
        require(
            float(target["properties"]["准时率%"]) == TARGET_NEW_ONTIME_RATE,
            "latest real dataset value was not projected into the ontology",
        )
        require(
            abs(float((target.get("computed") or {})["risk_score"]) - 13.6)
            < 0.000001,
            f"production derived value did not match the trial contract: {target}",
        )

        action_alerts = [
            row for row in updated_alerts
            if row.get("source") == "action"
            and (row.get("properties") or {}).get("status") == "open"
            and (row.get("properties") or {}).get("supplier_name")
            == target_supplier_name
        ]
        require(
            len(action_alerts) == 1,
            f"expected one real action-created alert, got {action_alerts}",
        )

        firings = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/firings?limit=100",
        )
        fired = [item for item in firings if item.get("status") == "fired"]
        fired_ids = {item.get("sentinelId") for item in fired}
        require(
            risk_sentinel_id in fired_ids,
            "risk Sentinel did not enter and execute its action",
        )
        require(
            alert_sentinel_id in fired_ids,
            "downstream Alert Sentinel was not triggered by the created object",
        )
        require(
            dynamic_sentinel_id in fired_ids,
            "assistant-created dynamic Sentinel did not execute on real data",
        )
        require(
            trial_probe_sentinel_id in fired_ids,
            "explicitly enabled built-in Sentinel did not execute its "
            "trial-proven existing match",
        )
        for item in fired:
            if item.get("sentinelId") not in {
                risk_sentinel_id,
                trial_probe_sentinel_id,
                alert_sentinel_id,
                dynamic_sentinel_id,
            }:
                continue
            results = item.get("actionResults") or []
            require(results, f"firing {item['id']} has no action results")
            require(
                all(result.get("status") == "success" for result in results),
                f"firing {item['id']} reported a failed action: {results}",
            )

        notifications = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/notifications?limit=100",
        )
        target_notifications = [
            item for item in notifications
            if item.get("actionId") == notify_action_id
            and target_supplier_name in str(item.get("body") or "")
        ]
        dynamic_notifications = [
            item for item in notifications
            if item.get("actionId") == dynamic_notify_action_id
            and target_supplier_name in str(item.get("body") or "")
        ]
        probe_notifications = [
            item for item in notifications
            if item.get("actionId") == notify_action_id
            and probe_supplier_name in str(item.get("body") or "")
        ]
        require(
            len(target_notifications) == 1,
            f"expected one durable notification, got {target_notifications}",
        )
        require(
            len(probe_notifications) == 1,
            "built-in activation did not produce exactly one downstream "
            f"notification: {probe_notifications}",
        )
        require(
            len(dynamic_notifications) == 1,
            "expected one assistant Sentinel notification, got "
            f"{dynamic_notifications}",
        )
        require(
            target_notifications[0].get("status") == "delivered",
            f"notification sink did not report delivery: {target_notifications[0]}",
        )
        require(
            dynamic_notifications[0].get("status") == "delivered",
            f"assistant notification sink did not report delivery: "
            f"{dynamic_notifications[0]}",
        )
        expected_notification_lineage = {
            notify_action_id: alert_sentinel_id,
            dynamic_notify_action_id: dynamic_sentinel_id,
        }
        for notification in (
            target_notifications + probe_notifications + dynamic_notifications
        ):
            require(
                notification.get("ontologyReleaseId") == release_id
                and notification.get("sentinelId")
                == expected_notification_lineage[notification["actionId"]]
                and bool(notification.get("actionLogId")),
                f"notification lineage is incomplete: {notification}",
            )

        action_logs = api.request(
            "GET",
            f"/api/v2/formal/ontologies/{ontology_id}/logs",
        )
        relevant_action_ids = {
            create_alert_action_id,
            notify_action_id,
            dynamic_notify_action_id,
        }
        relevant_logs = [
            item for item in action_logs
            if item.get("actionId") in relevant_action_ids
        ]
        successful_execution_logs = [
            item for item in relevant_logs if item.get("status") == "success"
        ]
        require(
            {item.get("actionId") for item in successful_execution_logs}
            == relevant_action_ids,
            "real actions are missing durable successful execution logs: "
            f"{relevant_logs}",
        )
        require(
            all(
                item.get("ontologyReleaseId") == release_id
                and bool(item.get("idempotencyKey"))
                and item.get("dryRun") is False
                for item in successful_execution_logs
            ),
            f"action execution lineage/idempotency is incomplete: "
            f"{successful_execution_logs}",
        )
        approved_proposals = [
            item for item in relevant_logs
            if item.get("id") == pending_create["id"]
        ]
        require(
            len(approved_proposals) == 1
            and approved_proposals[0].get("status") == "approved"
            and approved_proposals[0].get("relatedLogId")
            in {item.get("id") for item in successful_execution_logs},
            f"HITL decision/execution linkage is incomplete: {relevant_logs}",
        )

        fired_count = len(fired)
        notification_count = (
            len(target_notifications) + len(dynamic_notifications)
        )
        probe_notification_count = len(probe_notifications)
        action_alert_count = len(action_alerts)
        probe_action_alert_count = len([
            row for row in updated_alerts
            if row.get("source") == "action"
            and (row.get("properties") or {}).get("supplier_name")
            == probe_supplier_name
        ])
        require(
            probe_action_alert_count == 1,
            "built-in activation did not create exactly one real action object",
        )
        second_build = api.request(
            "POST",
            f"/api/v2/ontologies/{ontology_id}/mappings/build-all",
            json={},
        )
        assert_dispatch_success(second_build)
        final_cdc = wait_for_cdc(api, ontology_id)
        cdc_history = api.request(
            "GET",
            (
                f"/api/v1/ontologies/{ontology_id}/sentinels/cdc-status"
                f"?release_id={release_id}&include_history=true"
            ),
        )
        cdc_events = cdc_history.get("recent_events") or []
        require(
            cdc_events
            and all(
                event.get("ontologyReleaseId") == release_id
                and event.get("status") == "completed"
                for event in cdc_events
            )
            and any(
                int(event.get("cascadeDepth") or 0) >= 1
                for event in cdc_events
            ),
            f"durable CDC release/cascade lineage is incomplete: "
            f"{cdc_history}",
        )

        final_suppliers = api.request("GET", supplier_instances_path)
        final_alerts = api.request("GET", alert_instances_path)
        final_firings = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/firings?limit=100",
        )
        final_notifications = api.request(
            "GET",
            f"/api/v1/ontologies/{ontology_id}/sentinels/notifications?limit=100",
        )
        require(
            {
                key: row["id"]
                for key, row in by_property(
                    final_suppliers, "供应商ID").items()
            } == initial_supplier_ids,
            "repeated build changed stable Supplier IDs",
        )
        require(
            len([
                item for item in final_firings
                if item.get("status") == "fired"
            ]) == fired_count,
            "replaying the identical dataset produced duplicate fired edges",
        )
        require(
            len([
                item for item in final_notifications
                if item.get("actionId") in {
                    notify_action_id, dynamic_notify_action_id,
                }
                and target_supplier_name in str(item.get("body") or "")
            ]) == notification_count,
            "replaying the identical dataset produced a duplicate notification",
        )
        require(
            len([
                item for item in final_notifications
                if item.get("actionId") == notify_action_id
                and probe_supplier_name in str(item.get("body") or "")
            ]) == probe_notification_count,
            "replaying the identical dataset duplicated the built-in "
            "activation notification",
        )
        require(
            len([
                row for row in final_alerts
                if row.get("source") == "action"
                and (row.get("properties") or {}).get("supplier_name")
                == target_supplier_name
            ]) == action_alert_count,
            "replaying the identical dataset produced a duplicate action object",
        )
        require(
            len([
                row for row in final_alerts
                if row.get("source") == "action"
                and (row.get("properties") or {}).get("supplier_name")
                == probe_supplier_name
            ]) == probe_action_alert_count,
            "replaying the identical dataset duplicated the built-in "
            "activation action object",
        )

        return {
            "ontologyId": ontology_id,
            "releaseId": release_id,
            "sourceFile": str(data_file),
            "sourceRows": len(initial_supplier_ids),
            "targetSupplier": TARGET_SUPPLIER_ID,
            "oldConditionMatched": False,
            "newOnTimeRate": TARGET_NEW_ONTIME_RATE,
            "newRiskScore": 13.6,
            "stableSupplierIds": len(initial_supplier_ids),
            "stableAlertSeedIds": len(initial_seed_alert_ids),
            "actionCreatedAlerts": action_alert_count,
            "builtInActivationAlerts": probe_action_alert_count,
            "durableNotifications": notification_count,
            "builtInActivationNotifications": probe_notification_count,
            "hitlApprovalsExecuted": (
                len(approved_proposals) + len(probe_proposals)
            ),
            "dynamicSentinelId": dynamic_sentinel_id,
            "firedSentinels": sorted(str(item) for item in fired_ids),
            "repeatBuildDuplicateEffects": 0,
            "cdcHealthy": final_cdc.get("healthy"),
            "cdcQuiescent": final_cdc.get("quiescent"),
            "cdcCompleted": (
                cdc_history.get("durable") or {}
            ).get("completed"),
            "cdcMaxObservedCascadeDepth": max(
                int(item.get("cascadeDepth") or 0)
                for item in cdc_events
            ),
            "neo4jProjection": query_projection.get("neo4j"),
        }
    finally:
        api.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-url", default="http://127.0.0.1:8000")
    parser.add_argument("--username", default="admin")
    parser.add_argument("--password", default="admin123")
    parser.add_argument("--data-file", default=str(DEFAULT_DATA))
    return parser.parse_args()


def main() -> int:
    try:
        result = run(parse_args())
    except (CheckFailed, httpx.HTTPError, OSError, ValueError) as exc:
        print(f"REAL SENTINEL E2E FAILED: {exc}", file=sys.stderr)
        return 1
    print("REAL SENTINEL E2E PASSED")
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
